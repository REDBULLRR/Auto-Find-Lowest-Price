# encoding = utf-8
# File Name: shopClawers.py
# Last Changed: 2020.11.26
# Author: Amy
# 適用系統：Windows

# version: 4.0

# Important Note:
# This ver is FOR Windows ONLY.
#  
# This .py file stores the methods of Web Clawer on each required online shops
# Difference with 1st ver: use xlwings rather than openpyxl
# 自動查最低價.xlsm  file must be in the same directory as this python script

# Module used for crawling each business platform:
# selenium: momo
# requests: etmall(json), Find Price, Y購, Y城, PChome(json)

# 更動：
# 1. Y購爬蟲 Debug
# 2. color 從 自動查最低價.xlsm 裡面抓取（讓使用者可以自行設定）
# 3. 把可以設為 local 的 global 變數改為 local，加快執行速度 (ex: urls)
# 4. is_same_prod 加入 color 的判斷
# 5. getAmount_etmall 中的 UnboundLocalError: local variable 'amount' referenced before assignment 修復


# 急需優化：
# 1. 提高 3C 類產品的搜尋到的比率
# Ex: 
    #Apple Watch S5 GPS 44mm金色鋁金屬錶殼粉運動型錶帶
    #Apple Watch Series 5 GPS 44mm金色鋁金屬錶殼粉運動型錶帶
    #Apple iPad Air3 2019 LTE 64GG 10.5吋平板
    #Apple iPad Air3 2019 WIFI 64GG 10.5吋平板
    #Apple iPad7 2019 10.2吋 LTE 32G平板    
    #Apple iPad7 2019 10.2吋 WIFI 32G平板   Apple 2019 iPad 32G WiFi 10.2吋平板電腦
    #Samsung三星 Tab A (2019) 8吋 32G LTE平板(T295)

# 課題1:keyword making 
# Ex: (mapping to above)
    # Apple Watch S5 GPS 44mm
    # Apple Watch Series 5 GPS 44mm
    # Apple iPad Air3 2019 LTE 64GG 10.5
    # Apple iPad Air3 2019 WIFI 64GG 10.5
    # Apple iPad7 2019 10.2 LTE 32G
    # Apple iPad7 2019 10.2 WIFI 32G
    # Samsung Tab A (2019) 8 32G LTE(T295)

# 課題2：要能區分出商品與商品配件
# 例：
    # 預購 2019 iPad mini Wi-Fi+行動網路 256GB 7.9吋 平板電腦 晶豪泰3C 高雄 專業攝影
    # 
    # AMAZINGthing Apple iPad Mini 2019 (7.9") 專用繪圖保護膜

# 課題3: 要能判斷出相符品名
# Ex:
    # Samsung三星 Tab A (2019) 8吋 32G LTE平板(T295)
    # 【SAMSUNG 三星】Galaxy Tab A 2019 8吋 2G/32G LTE版 四核心平板電腦 SM-T295(送戎布保護套等多重好禮)
    #
    # Apple iPad7 2019 10.2吋 WIFI 32G平板
    # 【Apple 蘋果】2019 iPad 7 平板電腦(10.2吋/Wi-Fi/32G)
    # 
    # Apple iPad7 2019 10.2吋 WIFI 32G平板
    #【Apple 蘋果】2019 iPad 7 平板電腦(10.2吋/Wi-Fi/32G/贈Apple Pencil等好禮)
    
# idea:
    # 收集「辨別用關鍵字」：
    # mini, pro, wifi (wi-fi), lte

    # 商品わけ策：
    # removeChinese --> split into a list --> sort the list --> if the list are identical, send true

# 2. 搜尋不到規格完全吻合的商品時，容許使用者用 '____' 來代表未定的數字
# 已做，測試中

# 未來預計加入：
# 1. 加入 FP 確認商店是否有效（價錢正確與否、商品是否上架中）的確認機制
# 2. 讓使用者可以勾選要用多嚴謹的搜尋法：
#   非常嚴格     （工程師視角：規格必須吻合）
#   嚴格        （工程師視角：只有部份規格吻合。如：入數相同，盒數不同）
#   普通        （只要是同款商品即可，數量、規格等忽略）
# 3. 把品名中的空格符號用'*'取代掉，接著用 re.match 來比對兩者，如果有 match，那就接著檢查規格？
# （但可能會有非相同商品的例外狀況）

# 其他 Note：
# 老闆之前說的「讓手機關鍵字變得更簡潔」，其實想想，只要使用者輸入的關鍵字簡潔一點不就解決了嗎？
# MOMO 的最低價抓取精度尚未測試

from difflib import SequenceMatcher
from bs4 import BeautifulSoup as bs
from openpyxl import load_workbook
from urllib.parse import quote
import xlwings as xw
import platform,requests,logging,random, json,time,os,re

from selenium.webdriver.chrome.options import Options # For Headless mode
from selenium.webdriver.common.keys import Keys
from selenium import webdriver

#  - - - For Waits - - -
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By

# ======== Global Variables =========
# - - - 如果一直用同一個User Agent去請求網站回應的話，有可能會被鎖IP，所以這邊先列出一系列隨機的備用 UserAgent - - -
# 可新增。至 https://www.toolnb.com/tools-lang-zh-TW/createuseragent.html 獲取隨機 User Agent
UserAgents = [
    'Mozilla/5.0 (Macintosh; PPC Mac OS X 10_10_9 rv:2.0; he-IL) AppleWebKit/533.12.3 (KHTML, like Gecko) Version/5.0 Safari/533.12.3',
    'Mozilla/5.0 (iPad; CPU iPad OS 5_1_1 like Mac OS X) AppleWebKit/531.0 (KHTML, like Gecko) CriOS/22.0.834.0 Mobile/87E842 Safari/531.0',
    'Mozilla/5.0 (iPad; CPU iPad OS 5_1_1 like Mac OS X) AppleWebKit/531.0 (KHTML, like Gecko) CriOS/43.0.882.0 Mobile/37C008 Safari/531.0',
    'Mozilla/5.0 (iPod; U; CPU iPhone OS 3_3 like Mac OS X; uk-UA) AppleWebKit/531.31.5 (KHTML, like Gecko) Version/4.0.5 Mobile/8B117 Safari/6531.31.5',
    'Mozilla/5.0 (Macintosh; U; PPC Mac OS X 10_11_2 rv:5.0; dz-BT) AppleWebKit/532.45.2 (KHTML, like Gecko) Version/4.0 Safari/532.45.2',
    'Mozilla/5.0 (iPhone; CPU iPhone OS 6_1_6 like Mac OS X) AppleWebKit/532.2 (KHTML, like Gecko) FxiOS/10.9n4610.0 Mobile/30N900 Safari/532.2',
    'Mozilla/5.0 (iPod; U; CPU iPhone OS 3_0 like Mac OS X; gu-IN) AppleWebKit/533.20.7 (KHTML, like Gecko) Version/3.0.5 Mobile/8B115 Safari/6533.20.7']
headers = {
            'cookie': 'ECC=GoogleBot',
            # - - 使用隨機的 User-Agent - - - 
            "User-Agent":UserAgents[(random.randint(0,len(UserAgents)-1))]
            }

'''
result 的資料結構：
{
    'prod1': {}, <--- empty case

    'prod2':{
        'shopA':{
            'lowest_price': 19990,
            'foundName': 'SAMSUNG Galaxy A71 5G (8G/128G) 6.7吋智慧手機', 
            'imgUrl': 'https//media.etmall.com.tw/NXimg/002695/2695289/2695289_LM.jpg?t=17846951525', <-- if there is one
            'prodUrl': 'https://www.etmall.com.tw/SAMSUNG-Galaxy-A71-5G-8G-128G-6-7吋智慧手機/i/2695289'
            }
            }, 

    'prod3': {}, 
    Ex:    
    'Jo Malone 英國梨與小蒼蘭 香水 100ml': {
        'etmall': {
            'lowest_price': 4249, 
            'foundName': 'JO MALONE 英國梨與小蒼蘭香水100ml', 
            'imgUrl': 'https//media.etmall.com.tw/NXimg/002344/2344518/2344518_LM.jpg?t=17527103847', 
            'prodUrl': 'https://www.etmall.com.tw/JO-MALONE-英國梨與小蒼蘭香水100ml/i/2344518'}}, 
    
    'EBI ELIE SAAB 夢幻花嫁淡香精 TESTER 90ml': {}, 
    
    'MONTBLANC 萬寶龍 海洋之心女性淡香水 30ml 試用品TESTER': {}, 

    '豐力富 紐西蘭頂級純濃奶粉 2.6 公斤': {
        'etmall': {
            'lowest_price': 1051, 
            'foundName': '豐力富 紐西蘭頂級純濃奶粉 2.6 公斤', 
            'imgUrl': 'https//media.etmall.com.tw/NXimg/002708/2708912/2708912_LM.jpg?t=17856056486', 
            'prodUrl': 'https://www.etmall.com.tw/豐力富-紐西蘭頂級純濃奶粉-2-6-公斤/i/2708912'}}}
'''

'''
FindPrice 的搜尋結果特色：
1) 結果預設是由價格低到高排序。
2) 有時搜尋結果的第一項會是一個價格區間，點進去會有更多關鍵字符合的商品。
3) 那個價格區間的最低價如果不是我們要的最低價，我們可以單純無視。
'''    
# ========== Sub Functions =========
def getProdList():
    try: # Mac
        wb = load_workbook(os.path.split(os.path.realpath(__file__))[0] +'/自動查最低價.xlsm')
    except: # windows
        wb = load_workbook('自動查最低價.xlsm')

    ws = wb.active
    '''讀取 Excel 檔取得商品名單'''
    return [cell.value for cell in list(ws.columns)[1][2:] if not cell.value==None]

def is_same_prod(prod, found, color, threeC):
    '''判斷兩個品名是否為同一商品。\n
    ＊不去特別分辨 試用品 或 非試用品'''
    logging.info('-----------------------\n比對：\n%s\n%s',prod,found)

    # 兩品名皆轉換為小寫： 排除大小寫差異
    prod = removeComment(prod).lower()
    found = removeComment(found).lower()

    ___exist = False # default. It will be True if there is '___' in prod
    if '___' in prod:
        ___exist = True

    # 移除標點符號 
    prod = removePunc(prod)
    found = removePunc(found)

    # 遇到需要比較相似度的地方，我會用 re.sub('___','',prod) 來移除 prod 中的 '___'
    if SequenceMatcher(None, re.sub('___','',prod), found).quick_ratio() == 1:
        logging.info('OOO: %s \n 和\n%s 吻合！',re.sub('___','',prod),found)
        return True

    elif SequenceMatcher(None, re.sub('___','',prod), found).quick_ratio() > 0.76:
        logging.info('找到相似度超過 0.76 的商品')
        logging.info('第一階段商品名處理：小寫化、移除標點符號、移除宣傳語')
        logging.info('prod=%s',prod)
        logging.info('found=%s',found)

        # 檢查「指定規格」是否一致
        for word in color: # set color 的內容由使用者指定 (boom_data.xlsx)
            if word in prod and word not in found:
                logging.info('XXX: 指定檢查的規格不一致')
                print('指定檢查的規格不一致\n')
                return False

        # 如果商品是 3C ，用另外的特殊函式去判斷是否為相同商品 (目標是所有 3C 都在這邊處理)
        for brand in threeC:
            if brand.lower() in prod:
                logging.info('商品和 3C 有關，套用 3C 專用規格比較法')
                return is_same_specifi(prod, found, ___exist=___exist)
        
        # 特別處理：split 後前兩個字串都不是中文 （防：英文字太多，會使中文字串的兩三字間的差異（規格）被忽略）
        # 我承認這裡怪怪的，應該可以改得更有效率
        if not is_chinese(prod.split()[0])  and len(prod.split())>1  and not is_chinese(prod.split()[1]):
            if SequenceMatcher(None, removeNoChinese(re.sub('___','',prod)), removeNoChinese(found)).quick_ratio() > 0.75:
                s1 = ''
                s2 = ''

                # 抽掉它們不是中文的部分，重新比較
                for c in removeNoChinese(re.sub('___','',prod)).split():
                    s1 +=c

                for c in removeNoChinese(found).split():
                    s2 +=c

                # 若是只拿中文的部分去互相比較相似度依然高，則再比對規格是否相符
                if SequenceMatcher(None, s1, s2).quick_ratio() > 0.75:
                    logging.info('pass: 只拿中文的部分去互相比較相似度依然高')
                    logging.info('prod_chineseOnly=%s',s1)
                    logging.info('found_chineseOnly=%s',s2)

                    pass

                else:
                    logging.info('XXX: 只拿中文的部分去互相比較，相似度不足 0.75')
                    return False

            else:
                logging.info('XXX: 相似度不足 0.75')
                return False
    
    # 預防：品名實際一樣，但因為有其中一方有英文版的品牌名，一方沒有，導致相似度過低
    # 對應：拿掉品名中的所有英文，重比相似度
    # 20201123 筆：Error 百出，3.4 版中直接拿掉了這一塊
    # 20201112 筆: 這個部分常常誤殺第一段子字串並非英文的正確搜尋結果，要修改（參照隔壁的 jupyter note)
    elif SequenceMatcher(None, re.sub('___','',prod), found).quick_ratio() > 0.5 and ((len(prod.split())>1 and prod.split()[0].isalpha() and not is_chinese(prod.split()[0])) or (len(found.split())>1 and found.split()[0].isalpha() and not is_chinese(found.split()[0]))):
        logging.info('其中一方品名.split() 的第一段子字串不是中文，拿掉重新比較')
        logging.info('prod=%s',prod)
        logging.info('found=%s',found)
        logging.info('相似度大於 0.5 ，取走品名開頭非中文字串，再重新比較一次')

        if not is_chinese(prod.split()[0]):
            return is_same_prod(prod[prod.index(' ')+1:], found, color) 

        else:
            return is_same_prod(prod, found[found.index(' ')+1:], color)


    # 有的時候查到的相符品名的末尾會被賣家硬塞很多關鍵字，無法通過相似度測驗
    elif len(found.split()) > 3*len(prod.split()):
        logging.info('查到品名的末尾被賣家硬塞了很多關鍵字，重新檢查')
        same_count = 0
        for word in found.split():
            if word in prod:
                same_count+=1
        
        # 如果 found 的子字串（以空格分隔）有超過 3 串都有出現在 prod 的話，接著檢查規格
        if same_count > 3:
            logging.info('pass: 子字串檢查通過')
            pass

        else:
            logging.info('XXX: same_count=%d',same_count)
            return False
        
    # 相似度太低，排除
    else:
        logging.info('XXX: 相似度太低，排除')
        return False


    # Case1: 沒有要用'___'來當作隨意數字標記
    if not ___exist:     # 未完成
        #  ---檢查品名中出現的數字和數字順序(規格)是否一致---
        # 數字：指1,2,3,...。不包含one或一二三這種。
        numIn = ''
        numFoun = ''

        for n in [num if num.isdigit() else ' ' for num in prod ]:
            numIn+=n
        
        for n in [num if num.isdigit() else ' ' for num in found ]:
            numFoun+=n

        # 如果品名中出現的 數字 與 數字順序 不一致，則判斷為不同商品
        if  numIn.split() == numFoun.split():
            # 不過濾試用品
            logging.info('OOO: 規格相符\nnumIn=%s\nnumFoun=%s',numIn,numFoun)
            return True

        else:
            logging.info('numIn=%s',numIn)
            logging.info('numFoun=%s',numFoun)

            print('查到：',found)
            print('規格不符\n')

            logging.info('XXX: 品名中出現的 數字 與 數字順序 不一致，判斷為不同商品')
            return False

    # Case2: 有要用'___'來當作隨意數字標記
    else:
        #  ---檢查品名中出現的數字和數字順序(規格)是否一致: 製作正則表達式---
        # 數字：指1,2,3,...。不包含one或一二三這種。
        numIn = ''
        numFoun = ''

        for i in range(len(prod)):
            if prod[i].isdigit():
                numIn+= prod[i]

            elif i<len(prod)-2 and prod[i]=='_' and prod[i+1]=='_' and prod[i+2]:
                numIn+= '\d*'
    
        for n in found:
            if n.isdigit():
                numFoun+= n
        
        # 如果正則表達式檢查通過 (i.e., re.search() 回傳非 None 值），則回傳 True
        if re.search(numIn, numFoun):
            logging.info('OOO:\n搜尋：%s\n查到：%s\n正則表達式檢查通過！',prod,found)
            return True
        
        else:
            logging.info('XXX: 正則表達式檢查未通過\nnumIn=%s\nnumFoun=%s',numIn,numFoun)
            return False
        
def removeComment(astr):
    '''用正則表達式移除對搜尋沒太大幫助的宣傳語、註記'''
    
    # unicode編碼參考： https://zh.wikipedia.org/wiki/Unicode%E5%AD%97%E7%AC%A6%E5%88%97%E8%A1%A8#%E5%9F%BA%E6%9C%AC%E6%8B%89%E4%B8%81%E5%AD%97%E6%AF%8D
    newstr_list = re.sub(r'[\[【(「（][^（「\[【(]*(新品上市|任選|效期|專案|與.+相容|免運|折後|限定|獨家\d+折|福利品|現折|限時|安裝|適用|點數[加倍]*回饋|[缺出現司櫃]貨|結帳|促銷)[^\]【）(」]*[\]】）)」]|[(]([^/ ]+/ *){1,}[^/]+[)]|效期[\W]*\d+[./]\d+[./]*\d*|\d(選|色擇)\d|.(折後.+元.|[一二兩三四五六七八九十]+色|([黑紅藍綠橙黃紫黑白金銀]/)+.|\w選\w色|只要.+起)|[^\u0020-\u0204\u4e00-\u9fa5]|[缺出現司櫃]貨[中]*|[^ ]*安裝[^ ]*|下架|[^ ]*配送',' ',astr,6).split()

    newstr = ''
    # 去除頭尾空白字元
    for word in newstr_list:
        newstr +=word + ' '
    
    return newstr[:-1]

    # 去除頭尾空白字元
    for word in newstr_list:
        newstr +=word + ' '
    
    return newstr[:-1]

def is_same_specifi(prod1, prod2, ___exist=False):
    '''判斷規格是否一致，和 is_same_prod 中的判斷法不同的是，數字出現的順序不必要一樣\n 
    ___exist: flag; True: 使用者 有 使用'___'替代任意數字，反之為 False\n
    '''
    
    # iPad 篩選區： # 未完成
    if 'ipad' in prod1:
        # 判斷點1: iPad+數字 (ex: iPad7)
        if re.search('ipad *\d+',prod1) and re.search('ipad *\d+',prod1).group(0) in prod2:
            pass
        
        # 判斷點2： 年份 (ex: 2019)
        elif re.search('20\d\d',prod1) and re.search('20\d\d',prod1).group(0) in prod2:
            pass
        
        # 若有發現其他判斷點，則以 elif 繼續新增

        # 不符合此二判斷點的，都不是相同商品
        else: 
            return False

        


    # 跳轉到這裡，商品英文已全切成小寫且移除非必要標點符號
    # 將商品中的中文全都抽取掉
    prod1 = removeChinese(prod1).split()
    prod2 = removeChinese(prod2).split()

    prod1_remove_allEng = [w for w in prod1 if not w.isalpha()]
    prod2_remove_allEng = [w for w in prod2 if not w.isalpha()]

	#print(prod1_remove_allEng)
	#print(prod2_remove_allEng)

    return SequenceMatcher(None,prod1_remove_allEng,prod2_remove_allEng).quick_ratio()==1

def is_chinese(ustr):
    '''判斷一個字串是否全為中文字。 ＊只接受字串，不要亂丟int或其他什麼進來＊'''
    # 註：此處的中文字理論上包含所有繁體與簡體字
    # 只要字串包含以下任何一種，就會回傳 False ：全形標點符號、空白字元、英文、數字
    for uchar in ustr:
        if not(uchar >= u'\u4e00' and uchar <= u'\u9fa5'):
            return False

    return True

def removeNoChinese(astr):
    '''只保留原字串的中文與空白字元'''
    new_str = ''
    for c in astr:
        if is_chinese(c) or c in ' ':
            new_str+=c
        else:
            new_str+=' '

    return new_str

def removeChinese(astr):
    '''移除字串中所有中文字元'''
    new_str=''
    for c in astr:
        if not is_chinese(c):
            new_str+=c

    return new_str

def removePunc(astr):
    '''移除對商品名稱而言多餘（拿掉也不應該影響搜尋）的標點符號'''
    result = ''
    strl = list(astr)

    # 不用拿掉 + * \' .
    while set(strl).intersection({'》','《','「','」','【','】','!', '"', '#', '$', '%', '&', "'", '(', ')', ',', '-', '/', ':', ';', '<', '=', '>', '?', '@', '[', '\\', ']', '^',  '`', '{', '|', '}', '~'})!=set():
        i = 0
        for c in strl:
            if c in '《》!"#$%,;<=>?@\\^`|': # 把這些符號移除
                strl.remove(c)

            elif c in '&()[]{}:-~/「」【】\'': # 把這些符號替換成空白
                strl.insert(i,' ')
                strl.remove(c)

            i+=1

    # 把新字串拼回來
    for c in strl:
        result+=c

    return result

def only_leave_alnumblank(astr):
    '''把一個字串中的數字、英文、空白符號以外的字元都移除'''
    newstr = ''

    for c in astr:
        if c.isalnum() or c.isspace():
            newstr = newstr + c
    
    return newstr

def removeComma_and_toInt(str_list):
    '''移除一個由數字字串組成的 list/ str 中的每個字串中的標點符號(',','$','~','～')，並轉換字串為 int type'''
    # Last Chnaged: 2020.9.27

    # --- 處理傳入的變數並不是 list 類型的狀況 ---
    if type(str_list)==str: # str 
        astr = str_list

        while ',' in astr:
            astr = astr[:astr.find(',')]+astr[astr.find(',')+1:]

        while '$' in astr:
            astr = astr[:astr.find('$')]+astr[astr.find('$')+1:]
        
        while '~' in astr:
            astr = astr[:astr.find('~')]+astr[astr.find('~')+1:]

        while '～' in astr:
            astr = astr[:astr.find('～')] # 取左邊的

        while '元' in astr:
            astr = astr[:astr.find('元')]

        return int(astr)
        
    elif type(str_list)==int: # int
        return str_list     # 你幹嘛送不需要變換的東西進來啦
    # ---------------------------------------

    # --- 剩下的情形，預期：變數類型是 list ---
    if str_list==None:
        return None
    
    newlist = []
    if str_list==[]:
        pass

    else:
        for astr in str_list:#2020.9.1 更新
            if type(astr)==int:
                newlist.append(astr)
                continue

            while ',' in astr:
                astr = astr[:astr.find(',')]+astr[astr.find(',')+1:]

            while '$' in astr:
                astr = astr[:astr.find('$')]+astr[astr.find('$')+1:]
            
            while '~' in astr:
                astr = astr[:astr.find('~')]+astr[astr.find('~')+1:]

            while '～' in astr:
                astr = astr[:astr.find('～')] # 取左邊的

            while '元' in astr:
                astr = astr[:astr.find('元')]

            newlist.append(int(astr))    

    return newlist

# ========= 賣場可賣量抓取/ 價錢比對(isUrlAvailiable) ===========
def getAmount_PC(urls):
    '''使用 selenium 抓取 PChome 商品頁面的最大可賣量\n
    urls 是裝有網址的字串 list，回傳一個正整數的 list '''
    logging.info('Start getAmount_PC(urls)')
    
    # 無頭模式 報錯: 原始碼：loading 網頁載入中
    # 不知怎樣可解決此問題
    chrome_options = Options()
    chrome_options.add_argument('--no-sandbox') # 解決 webdriver automation extension 打不開的問題
    chrome_options.add_argument('user-agent='+UserAgents[random.randint(0,len(UserAgents)-1)]) # 解決抓到的原始碼是 loading 網頁載入中 的問題

    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')

    driver = webdriver.Chrome(os.path.split(os.path.realpath(__file__))[0]+r'\chromedriver.exe',options = chrome_options)
    #wait = WebDriverWait(driver, 20)

    amounts = []
    j = 1
    for url in urls:
        amount = '-'

        if url=='-':
            amounts.append('-')
            continue
        
        driver.get(url)
        driver.implicitly_wait(30)
        #wait.until(EC.presence_of_element_located((By.CLASS_NAME,'Qty'))) 去掉這行好像就不會有視窗載不出來的問題
        soup = bs(driver.page_source, 'html.parser') 
        #try:

        # Case1: 網頁顯示：商品售完補貨中
        if soup.find('li',id='ButtonContainer') is not None and '售完，補貨中！' in soup.find('li',id='ButtonContainer').text:
            amount = '售完補貨中'

        # Case2: 網頁顯示：商品完售
        elif soup.find('li',id='ButtonContainer') is not None and '完售，請參考其他商品' in soup.find('li',id='ButtonContainer').text:
            amount = '完售'

        # Case3: 以上兩者都沒發生：預期商品頁面有顯示商品可賣量
        else:
            try:
                amount = int(soup.find('select', class_='Qty').find_all('option')[-1].text)

            except AttributeError:
                print("soup.select('select'):",soup.select('select'))
                print("soup.select('button')",soup.select('button'))
                #print("soup.find('select', class_='Qty').find_all('option')[-1].text=",soup.find('select', class_='Qty').find_all('option')[-1].text) # AttributeError: 'NoneType' object has no attribute 'find_all'
                amount = 'AttributeError'
                time.sleep(30)

        '''except: # 完售
            print('error')
            print()
            amount = 'error'''
        amounts.append(amount)

        print('可賣量查詢進度：{}/{}'.format(j,len(urls)))
        j+=1

    driver.quit()
    return amounts

def getAmount_momo(driver, wait, url):
    '''取得 momo 商品網址內的商品可賣量\n
    一次只讀入一個網址'''
    logging.info('Start getAmount_momo(driver, wait, url)')

    driver.get(url)
    wait.until(EC.presence_of_element_located((By.CLASS_NAME,'CompareSel')))
    soup = bs(driver.page_source,'html.parser')
    amount = int(soup.find('select', class_='CompareSel', id= 'count').find_all('option')[-1].text)

    return amount
  
def getAmount_etmall(urls):
    '''使用 selenium 抓取 東森購物 商品頁面的最大可賣量\n
    urls 是裝有網址的字串 list，回傳一個正整數的 list 。'''
    logging.info('Start getAmount_etmall(urls)')

    chrome_options = Options()
    chrome_options.add_argument('--no-sandbox') # 解決 webdriver automation extension 打不開的問題
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    #chrome_options.add_argument('user-agent='+UserAgents[random.randint(0,len(UserAgents)-1)])

    # ---- 開啟 chrome driver  -----
    if platform.system() == 'Windows': # Windows
        driver = webdriver.Chrome(os.path.split(os.path.realpath(__file__))[0]+r'\chromedriver.exe',options = chrome_options)

    elif platform.system() == 'Darwin': # Mac OS
        driver = webdriver.Chrome('./chromedriver',options = chrome_options)
    # ---- chrome driver 準備完畢 ----

    amounts = []
    j=1
    for url in urls:
        if url=='-':
            amounts.append('-')
            continue
        driver.get(url)
        driver.implicitly_wait(5)
        soup = bs(driver.page_source, 'html.parser') 
        amount = 'No_Found'

        try:
            if soup.find('select', class_='t-quantitySelector n-form--control') is not None:
                amount = int(soup.find('select', class_='t-quantitySelector n-form--control').find_all('option')[-1].text)

            elif soup.find('select', class_='t-quantitySelector') is not None:
                amount =  int(soup.find('select', class_='t-quantitySelector').find_all('option')[-1].text)

        except AttributeError:
            print('AttributeError')

            if soup.find('a',class_='n-btn n-btn--disabled') is not None and '銷售一空'  in soup.find('a',class_='n-btn n-btn--disabled').text:
                amount = '售完'

            else:
                print('未知錯誤，印出原始頁面內容：')  
                print(soup.text)
                time.sleep(20)
                amount = 'error'
            
        amounts.append(amount)
        print('進度：{}/{}'.format(j,len(urls)))
        j+=1

    driver.quit()
    return amounts  

def getStatus_Ymall(urls):
    '''抓取商品們處於可買或已售完狀態。回傳 list'''
    logging.info('Start getStatus_Ymall(urls)')

    s = requests.session()
    status = []

    for url in urls:
        req = s.get(url, headers= headers)
        soup = bs(req.text, 'html.parser')
        if '售完補貨中' in soup.find('div',id='ypsa2ct-2015').text:
            status.append('售完')

        elif '立即購買' in soup.find('div',id='ypsa2ct-2015').text:
            status.append('可買')

        else: # 我也很好奇這會是什麼狀況
            status.append('兩者皆非')
            print(soup.find('div',id='ypsa2ct-2015').text)
        
        time.sleep(random.randint(1,10)*0.1)
    
    return status

def isUrlAvailiable(url, shop, price):
    '''檢查在 FindPrice 上所查到的商品價格是否正確、店面是否存在'''
    logging.info('Start isUrlAvailiable(url, shop, price)')

    r = requests.get(url, headers=headers)
    r.encoding = 'utf-8'
    soup = bs(r.text,'html.parser')

    if shop == '樂天市場':
        try:
            return removeComma_and_toInt(soup.select_one('#auto_show_prime_price > strong > span').text) == price
            
        except:
            print(url)
            if soup.select_one('#auto_show_prime_price > strong > span')==None:
                print('[樂天市場] 找不到價格欄\n')
            else:
                print('[樂天市場] 價錢和Find Price 所顯示的不符\n')
            
            return False

    elif shop =='Yahoo奇摩超級商城':
        try:
            return soup.find('span',class_='price').text[:-1] == str(price)
        except:
            print(url)
            if soup.find('span',class_='price')==None:
                print('[Yahoo奇摩超級商城] 找不到價格欄\n')
            else:
                print('[Yahoo奇摩超級商城] 價錢和Find Price 所顯示的不符\n')
            return False

    elif shop =='Yahoo奇摩購物中心':
        try: 
            return removeComma_and_toInt(soup.select_one('#isoredux-root > div > div.ProductItemPage__pageWrap___2CU8e > div > div:nth-child(1) > div.ProductItemPage__infoSection___3K0FH > div.ProductItemPage__rightInfoWrap___3FNQS > div > div.HeroInfo__heroInfo___1V1O8 > div > div.HeroInfo__leftWrap___3BJHV > div > div').text)==price
        except:
            print(url)
            if soup.select_one('#isoredux-root > div > div.ProductItemPage__pageWrap___2CU8e > div > div:nth-child(1) > div.ProductItemPage__infoSection___3K0FH > div.ProductItemPage__rightInfoWrap___3FNQS > div > div.HeroInfo__heroInfo___1V1O8 > div > div.HeroInfo__leftWrap___3BJHV > div > div')==None:
                print('[Yahoo奇摩購物中心] 找不到價格欄\n')
            else:
                print('[Yahoo奇摩購物中心] 價錢和Find Price 所顯示的不符\n')
            return False

    elif shop =='myfone購物':
        try: 
            return removeComma_and_toInt(soup.select_one('#item-419 > div.wrapper > div.section-2 > div.prod-description > div.prod-price > span.prod-sell-price').text) == price
            
        except:
            print(url)
            if soup.select_one('#item-419 > div.wrapper > div.section-2 > div.prod-description > div.prod-price > span.prod-sell-price')==None:
                print('[myfone購物] 找不到價格欄\n')
            else:
                print('[myfone購物]價錢和Find Price 所顯示的不符\n')
            
    elif shop == 'momo購物網':
        pass

    elif shop == 'PChome 24h購物':
        try:
            return soup.select_one('#PriceTotal').text==str(price)

        except:
            print(url)
            if soup.select_one('#PriceTotal')==None:
                print('[PChome 24h購物]找不到價格欄\n')
            else:
                print('[PChome 24h購物]價錢和Find Price 所顯示的不符\n')

            return False

    elif shop == 'ETmall東森購物網':
        try:
            return removeComma_and_toInt(soup.select_one('#productDetail > div:nth-child(2) > section > section > div:nth-child(3) > div.n-price__block > div.n-price__bottom > span.n-price__exlarge > span.n-price__num').text)==price
        except:
            print(url)
            if soup.select_one('#productDetail > div:nth-child(2) > section > section > div:nth-child(3) > div.n-price__block > div.n-price__bottom > span.n-price__exlarge > span.n-price__num')==None:
                print('[ETmall東森購物網] 找不到價格欄\n')
            else:
                print('[ETmall東森購物網] 價錢和Find Price 所顯示的不符\n')
                
            return False

    elif shop == '蝦皮商城':
        pass

    elif shop == 'Costco 好市多線上購物':
        pass
    elif shop == 'udn買東西':
        pass

    elif shop == 'friDay購物':
        try:
            return soup.select_one('#E3 > div > div > div.prodinfo_area > span > div.bayPricing_area > div.attract_block > span.useCash > span.price_txt').text == str(price)
        except:
            print(url)
            if soup.select_one('#E3 > div > div > div.prodinfo_area > span > div.bayPricing_area > div.attract_block > span.useCash > span.price_txt')==None:
                print('[friDay購物] 找不到價格欄\n')
            else:
                print('[friDay購物] 價錢和Find Price 所顯示的不符\n')
            return False

    elif shop == 'PChome 商店街':
        pass
    elif shop == '家樂福線上購物網':
        pass
    elif shop == 'momo摩天商城':
        try:
            return removeComma_and_toInt(soup.select_one('#goodsForm > div.prdInnerArea > div > div.prdrightwrap > div.prdleftArea > div.prdDetailedArea > dl > dd.sellingPrice > span').text) == price
        except:
            print(url)
            if soup.select_one('#goodsForm > div.prdInnerArea > div > div.prdrightwrap > div.prdleftArea > div.prdDetailedArea > dl > dd.sellingPrice > span')==None:
                print('[momo摩天商城] 找不到價格欄\n')
            else:
                print('[momo摩天商城] 價錢和Find Price 所顯示的不符\n')

    elif shop == '森森購物網':
        pass
    elif shop == '愛買線上購物':
        pass
    elif shop == '大樹健康購物網':
        pass
    elif shop == '小三美日平價美妝':
        pass
    elif shop == '松果購物':
        pass
    elif shop == '生活市集':
        pass
    elif shop == 'Jollybuy 有閑':
        pass
    elif shop == '熊媽媽買菜網':
        pass
    elif shop == '淘寶精選':
        pass

'''def getAmount_FP(urls, shops):
'''
# ================================
# ========== Crawlers ============

def crawler_on_FP(prod_list, ws, color, threeC):
    '''抓取比價網站 Find Price 上的最低價'''
    logging.info('Crawl on FP')

    fp_mainurl = 'https://www.findprice.com.tw/g/'
    fp_mainurl2 = 'https://www.findprice.com.tw'

    # 印分隔線
    print('*'*15+' Find Price '+'*'*15)
    result_for_write = []
    s = requests.session()

    j=0
    try:
        for prod in prod_list:
            j+=1
            ws['a1'].value = 'FP：{0}/{1}'.format(j,len(prod_list))
            logging.info('搜尋商品：%s'%re.sub('___',' ',prod))
            print('搜尋商品：{}\n'.format(re.sub('___',' ',prod)))
            #result[prod]['FP'] = {}
            names = []
            urls = []
            prices = []
            shops = []

            url = fp_mainurl + quote(re.sub('___',' ',prod))
            req = s.get(url, headers=headers)
            count =1
            while not req.ok:
                logging.warning('連線 Find Price 失敗，請檢查網路狀態。程式接下來每隔三秒會重新嘗試連線')
                print('連線 Find Price 失敗，請檢查網路狀態。程式接下來每隔三秒會重新嘗試連線')
                time.sleep(3)

                req = s.get(url, headers=headers)
                count+=1
                if count>5:
                    logging.error('連線失敗次數過多，自動跳出 Find Price')
                    raise TimeoutError('連線失敗次數過多，自動跳出 Find Price')

            req.encoding = 'utf-8'

            soup = bs(req.text ,'html.parser')

            if soup.find('div',id='GoodsGridDiv') == None or soup.find('div',id='GoodsGridDiv').a == None:
                # 有時候會出現：所有商品（包括人工搜時有搜尋結果的商品）都沒有搜尋結果的狀況
                # 原因應該是頁面抓不到這個 if 上面在判斷的 div tag
                logging.debug("沒有抓到 id='GoodsGridDiv' 的 div 標籤")
                logging.debug('req.json = %s'%req.json) 
                logging.debug("[div.text for div in soup.find_all('div')] = %a"%[div.text for div in soup.find_all('div')])

                print('無任何搜尋結果')
                result_for_write.append(['-','-','-','-'])

                print('='*35)
                continue

            # 麻煩的傢伙來了
            if soup.find('div', id='HotDiv').table != None:
                logging.info('FP: 麻煩的搜尋結果出現了，請前方人員進入戰鬥狀態')
                
                link = fp_mainurl2 + soup.find('div', id='HotDiv').a['href']
                reqq = s.get(link, headers=headers)
                reqq.encoding = 'utf-8'
                soup2 = bs(reqq.text, 'html.parser')

                found_names = [tr.find_all('td')[2].a.text for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

                found_prices = [tr.find_all('td')[1].text for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

                found_urls=[tr.find_all('td')[2].a['href'] for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

                found_shops = [tr.find_all('td')[2].img['title'] for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

                found_prices = removeComma_and_toInt(found_prices)

                logging.debug('唧')

                for i in range(len(found_names)):
                    if is_same_prod(prod, found_names[i], color,threeC):
                        logging.info('查到：%s'%found_names[i])
                        print('查到：',found_names[i])
                        names.append(found_names[i])
                        prices.append(found_prices[i])
                        urls.append(fp_mainurl2 + found_urls[i])
                        shops.append(found_shops[i])
                        
                        logging.debug('呱')

            # 一般來說會直接用這邊的程式碼
            found_names = [tr.find_all('td')[1].a.text for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')]

            found_prices = [tr.find_all('td')[1].span.text for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')]

            found_urls=[tr.find_all('td')[1].a['href'] for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')]

            found_shops = [tr.find_all('td')[1].img['title'] for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')]

            try:
                for i in range(len(found_names)):
                    if is_same_prod(prod, found_names[i], color,threeC):
                        logging.info('查到：%s'%found_names[i])
                        print('查到：',found_names[i])
                        names.append(found_names[i])
                        prices.append(found_prices[i])
                        urls.append(fp_mainurl2 + found_urls[i])
                        shops.append(found_shops[i])

            except AttributeError: # found_names[i] 是 list?
                logging.error('FP crawler: AttributeError')
                logging.info('names=',names)
                logging.info('prices=',prices)
                logging.info('urls=',urls)
                logging.info('shops=',shops)

            prices = removeComma_and_toInt(prices)

            if names == []:
                logging.info('無相符之搜尋結果')
                print('無相符之搜尋結果')
                print('='*35)

                result_for_write.append(['-','-','-','-'])
                continue
            
            logging.info('抓到相符商品數：%d\n'%len(names))
            print('抓到相符商品數：',len(names),'\n')
            # 存下資料以待寫入
            result_for_write.append([min(prices), names[prices.index(min(prices))], shops[prices.index(min(prices))], urls[prices.index(min(prices))]]) # 不同平台的商店所對應的可賣量位置不同，待補。現階段只輸出網址
            print('='*35)

    except:
        logging.exception('Error occour when crawling Find Price')

    finally:    
        # 將資料寫入儲存格
        ws['u3'].value = result_for_write
        logging.info('Find Price 抓取完畢！\n')
        print('Find Price 抓取完畢！\n')

def crawler_on_Ymall(prod_list,ws, color, threeC):
    '''抓取 Y城 上的最低價'''
    logging.info('Crawl on Ymall')
    Ym_url = 'https://tw.search.mall.yahoo.com/search/mall/product?'

    # 印分隔線
    print('*'*15+' Yahoo!超級商城 '+'*'*15)
    result_for_write = []

    s = requests.session()

    j=0
    try:
        for prod in prod_list:
            j+=1
            ws['a1'].value = 'Y城：{0}/{1}'.format(j,len(prod_list))
            logging.info('搜尋商品：%s'%re.sub('___',' ',prod))
            print('搜尋商品：{}\n'.format(re.sub('___',' ',prod)))
            #result[prod]['Ymall'] = {}
            names = []
            urls = []
            prices = []

            req = s.get(Ym_url, params={'p':re.sub('___',' ',prod),'qt':'product','sort':'p'}, headers=headers)
            req.encoding = 'utf-8'
            logging.info('%s',req.url)

            soup = bs(req.text ,'html.parser')

            if  soup.find('ul',class_='gridList')== None:
                logging.info("soup.find('ul',class_='gridList')== None, 無任何搜尋結果")

                print('無任何搜尋結果')
                print('='*35)
                result_for_write.append(['-','-','-'])
                continue
            
            # ----- 抓取搜尋結果網頁上的品名、價格、網址 ------
            try:
                found_names = [x.string for x in soup.find('ul',class_='gridList').find_all('span', class_='BaseGridItem__title___2HWui')] 

                found_prices = [x.string[1:] for x in soup.find('ul',class_='gridList').find_all('em')]

                found_urls = [li.a['href'] for li in soup.find('ul',class_='gridList').find_all('li',class_='BaseGridItem__grid___2wuJ7')]


            except TypeError: # type 轉換時的錯誤預防
                logging.warning('Ymall: TypeError')

                for li in soup.find('ul',class_='gridList').find_all('li'):
                    logging.info("搜尋到的商品列表：\nfor li in soup.find('ul',class_='gridList').find_all('li'):\nli.text=%s",li.text)

            if not found_prices or not found_urls:
                logging.warning('Search XXX: 抓到的商品價格或商品網址為空。目標網站結構可能有變，原始碼需要修改')
                logging.warning('抓到商品價格：%a',found_prices)
                logging.warning('抓到商品網址：%a',found_urls)
                print('目標網頁結構可能有變，沒有抓到目標網址或價格')
                print('跳過 Y城 的搜尋')
                break

            # 過濾並存取相符品名
            for i in range(len(found_names)):
                if is_same_prod(prod, found_names[i],color,threeC):
                    logging.info('查到：%s',found_names[i])
                    print('查到：',found_names[i])
                    names.append(found_names[i])
                    prices.append(found_prices[i])
                    urls.append(found_urls[i])

            prices = removeComma_and_toInt(prices)

            if names == []:
                logging.info('無相符之搜尋結果')
                print('無相符之搜尋結果')
                print('='*35)
                result_for_write.append(['-','-','-'])
                continue

            logging.info('抓到相符商品數：%d',len(names))
            logging.info('抓到商品：%a',names)
            print('抓到相符商品數：',len(names),'\n')
            # 將資料存入以待寫入
            result_for_write.append( [min(prices), names[prices.index(min(prices))], urls[prices.index(min(prices))]])

            ##result[prod]['Ymall']['prodUrl'] = urls[prices.index(min(prices))]
            print('='*35)

    except:
        logging.exception('Error in Ymall!')

    finally:            
        # 將資料寫入儲存格
        ws['c3'].value = result_for_write
        logging.info('Yahoo!超級商城 抓取完畢！\n')
        print('Yahoo!超級商城 抓取完畢！\n')

def crawler_on_Ybuy(prod_list, ws, color, threeC):
    '''抓取 Y購 上的最低價'''
    logging.info('Crawl on Ybuy')

    # 問題：搜尋不到商品
    # 策略：拆為關鍵字搜尋

    Yb_url = 'https://tw.buy.yahoo.com/search/product?'
    Yb_official_url = 'https://tw.buy.yahoo.com/?sub=283'

    # 印分隔線
    print('*'*15+' Yahoo!購物中心 '+'*'*15)
    result_for_write = []
    s = requests.session()

    j=0
    try:
        for prod in prod_list:
            j+=1
            ws['a1'].value = 'Y購：{}/{}'.format(j,len(prod_list))
            logging.info('搜尋商品：%s'%re.sub('___',' ',prod))
            print('搜尋商品：{}\n'.format(re.sub('___',' ',prod)))
            #result[prod]['Ybuy'] = {}
            names = []
            urls = []
            prices = []

            req = s.get(Yb_url, params={'p':re.sub('___',' ',prod),'sort':'price'}, headers=headers)
            req.encoding = 'utf-8'
            # test
            #ws['a1'].value = req.url

            soup = bs(req.text ,'html.parser')

            if soup.find('ul',class_='gridList')== None:
                # Y拍賣上有相似商品：改去Y拍上面搜尋
                if soup.select_one('#isoredux-root > div.page.shopping > div > div > div > div > div:nth-child(3) > div > span.NoResult_title_2bt6D'):
                    req = s.get()

                else:
                    logging.info('Ybuy: 無任何搜尋結果')
                    print('無任何搜尋結果')
                    print('='*35)
                    result_for_write.append(['-','-','-'])
                    continue
            
            found_names = [ele.text for ele in soup.select('span.BaseGridItem__title___2HWui')]  # 爬取品名

            found_prices = [span.em.text for span in soup.select('span.BaseGridItem__itemInfo___3E5Bx')]# 爬取價格
            
            found_urls = [child.a['href'] for child in soup.find_all('ul',class_='gridList')[-1].children if child.a!=None and child.a['href']!=Yb_official_url] # 爬取商品商店網址

            # test
            #print('len(found_names):',len(found_names)) #5
            #print('len(found_prices):',len(found_prices))#4
            #time.sleep(3)


            '''這段不能用！這個只（？）抓得到旁邊有「劃掉的價錢」的價格搜尋結果'''
            #if len(found_prices) != len(found_names): # 有時候要用下面這行才抓得到價錢
            #    found_prices = [span.em.text for span in soup.select('span.BaseGridItem__price___31jkj')]            
            #    print('呱')
            #    print('len(found_prices):',len(found_prices))#1
                #print(soup.find_all('ul',class_='gridList')[-1].text)


            # 檢查抓到的商品價格與網址是否為空
            if not found_prices or not found_urls:
                logging.warning('Search XXX: 抓到的商品價格或商品網址為空。目標網站結構可能有變，原始碼需要修改')
                logging.warning('抓到商品價格：%a',found_prices)
                logging.warning('抓到商品網址：%a',found_urls)
                print('目標網頁結構可能有變，沒有抓到目標網址或價格')
                print('跳過 Y城 的搜尋')
                break
            
            # 檢查爬到的「商品名筆數」是否和「商品價格筆數」一致
            if len(found_names)!=len(found_prices):
                logging.warning('從原始碼中抓到的品名數和價格數不同')
                logging.info('prod:%s',prod)
                logging.info('found_prices:%a',found_prices)
                logging.info('found_names:%a',found_names)
                logging.info('found_urls:%a',found_urls)
                logging.info("soup.select('em.BaseGridItem__price___31jkj')=%a",soup.select('em.BaseGridItem__price___31jkj'))
            
            try:
                # 過濾並存取相符品名
                for i in range(len(found_names)):
                    if is_same_prod(prod, found_names[i], color,threeC):
                        logging.info('查到：%s'%found_names[i])
                        print('查到：',found_names[i])
                        names.append(found_names[i])
                        prices.append(found_prices[i])
                        urls.append(found_urls[i])

            except:
                logging.exception('Ybuy:過濾並存取相符品名時出現錯誤')
                logging.info('prod:%s',prod)
                logging.info('found_prices:%a',found_prices)
                logging.info('found_names:%a',found_names)
                logging.info('found_urls:%a',found_urls)
                logging.info('i:%d',i)
                logging.info("soup.select('em.BaseGridItem__price___31jkj')=%a",soup.select('em.BaseGridItem__price___31jkj'))

            prices = removeComma_and_toInt(prices)

            if names == []:
                logging.info('無相符之搜尋結果')
                print('無相符之搜尋結果')
                print('='*35)
                result_for_write.append(['-','-','-'])
                continue
            
            logging.info('抓到相符商品數：%d',len(names))
            logging.info('抓到商品：%a',names)
            print('抓到相符商品數：',len(names))

            # 將資料存入
            result_for_write.append([min(prices),names[prices.index(min(prices))], urls[prices.index(min(prices))]])
            ##result[prod]['Ybuy']['lowest_price'] = 
            ##result[prod]['Ybuy']['foundName'] = 
            ##result[prod]['Ybuy']['prodUrl'] = urls[prices.index(min(prices))]
            print('='*35)

    except:
        logging.exception('Error occour when crawling Ybuy!')
    
    finally:            
        # 將資料寫入儲存格
        ws['f3'].value = result_for_write

        print('Yahoo!購物中心 抓取完畢！\n')

def crawler_on_etmall(prod_list,ws, color, threeC):
    '''抓取 東森購物 上的最低價'''
    logging.info('Crawl on etmall')

    et_url = 'https://www.etmall.com.tw/Search/Get'   # 取得 json 用（從封包 GET 中取得)
    et_mainurl = 'https://www.etmall.com.tw'          # 拼接出商品頁面網址用

    # 印分隔線
    print('*'*15+' 東森購物網 '+'*'*15)
    result_for_write = []
    s = requests.Session()

    j=0
    try:
        for prod in prod_list:
            j+=1
            ws['a1'].value = ['東森：{0}/{1}'.format(j,len(prod_list))]
            #result[prod]['etmall'] = {}
            names = []
            urls = []
            image_urls = []
            prices = []

            page = 0
            payload = {
                'keyword':re.sub('___',' ',prod),
                'model[cateName]':'全站',
                'model[page]':0,
                'model[storeID]':'',
                'model[cateID]':-1,
                'model[filterType]':'',
                'model[sortType]':'',
                'model[moneyMaximum]':'',
                'model[moneyMinimum]':'',
                'model[pageSize]':'48',
                'model[SearchKeyword]':'',
                'model[fn]':'',
                'model[fa]':'',
                'model[token]':'',
                'model[bucketID]':1,
                'page':page
            }
            headers={
                "User-Agent":UserAgents[(random.randint(0,len(UserAgents)-1))]
                }
            req = s.post(et_url, data=payload, headers=headers) # 網頁會以 json 格式回傳表單搜尋結果

            if req.ok:
                print('搜尋商品：{}\n'.format(re.sub('___',' ',prod)))
            else:
                logging.warning('連線失敗')
                raise ConnectionError('連線失敗，狀態碼: {}，請檢查網路連線。\n'.format(req.status_code))

            req.encoding = 'utf-8'
            data = json.loads(req.text)

            # 無搜尋結果 --> 接著搜下一個商品  *改進計畫：考慮去過濾「相似商品」
            if data['searchResult']['isFuzzy']:
                logging.info("data['searchResult']['isFuzzy'] == False, 無搜尋結果")
                print('商品：{}\n無搜尋結果'.format(re.sub('___',' ',prod)))
                print('='*35)
                result_for_write.append(['-','-','-','-'])
                continue   
            
            # 有搜尋結果 --> 開始過濾
            else:
                print('一共有{}頁搜尋結果. . .'.format(data['searchResult']['totalPages']))
                print('一共有{}筆資料. . .\n'.format(data['searchResult']['totalProducts']))

                # - - - - 讀取每一頁的結果 - - - - - 
                for i in range(data['searchResult']['totalPages']):
                    logging.info('第 %d 頁'%(page+1))
                    print('- 第 {} 頁 -'.format(page+1))

                    # - - - 存取每頁的品名、賣價、商品連結、商品圖片網址 - - - 
                    for item_dict in data['searchResult']['products']:
                        name = item_dict['title']                   # 品名
                        link = et_mainurl+item_dict['pageLink']     # 商品頁面網址
                        imageUrl = 'https'+item_dict['imageUrl']    # 商品圖片網址
                        price = item_dict['finalPrice']             # 商品賣價

                        #print(is_same_prod(prod, name))
                        
                        # - - - 把和給定品名相符的品名資料存入 list - - -
                        if is_same_prod(prod, name, color,threeC):
                            logging.info('查到：%s'%name)
                            print('查到:{}'.format(name))
                            names.append(name)
                            urls.append(link)
                            image_urls.append(imageUrl)
                            prices.append(int(price))
                    
                    page = i+1
                    payload = {
                        'keyword':re.sub('___',' ',prod),
                        'model[cateName]':'全站',
                        'model[page]':0,
                        'model[storeID]':'',
                        'model[cateID]':-1,
                        'model[filterType]':'',
                        'model[sortType]':'',
                        'model[moneyMaximum]':'',
                        'model[moneyMinimum]':'',
                        'model[pageSize]':'48',
                        'model[SearchKeyword]':'',
                        'model[fn]':'',
                        'model[fa]':'',
                        'model[token]':'',
                        'model[bucketID]':1,
                        'page':page
                    }
                    req = s.post(et_url, data=payload, headers=headers)
                    req.encoding = 'utf-8'
                    data = json.loads(req.text) 
            
            if names == []:
                logging.info('name == [], 無相符搜尋結果')
                print('無相符搜尋結果')
                print('='*35)
                result_for_write.append(['-','-','-','-'])
                continue

            else:
                print('共有{}項相符搜尋結果'.format(len(names)))
                try:
                    result_for_write.append([min(prices),names[prices.index(min(prices))], '正在查詢', urls[prices.index(min(prices))]])

                except:
                    logging.exception('Error occour when crawling PChome!')
                    logging.info('names=%s',names)
                    logging.info('prices=%s',prices)
                    logging.info('urls=%s',urls)
                    logging.info('result_for_write=%a',result_for_write)
                
                ##result[prod]['etmall']['imgUrl'] = image_urls[prices.index(min(prices))]
                ##result[prod]['etmall']['prodUrl'] = urls[prices.index(min(prices))]
                print('='*35)
    
    except:
        logging.exception('Error occour when crawling PChome!')

    finally:
        # 將資料寫入儲存格
        ws['q3'].value = result_for_write
        logging.info('東森：可賣量查詢中')
        ws['a1'].value = '東森：可賣量查詢中'

        urls = [alist[-1] for alist in result_for_write]
        amounts = getAmount_etmall(urls)
        logging.info('amounts=%d',amounts)
        amounts_to_write = [[num] for num in amounts]
        ws['s3'].value = amounts_to_write

        logging.info('東森購物網抓取完畢！\n')
        print('東森購物網抓取完畢！\n')

def crawler_on_momo(prod_list,ws, color, threeC):
    '''用 selemium 上 momo 抓取商品最低價'''
    logging.info('Crawl on momo')

    momo_url ='https://www.momoshop.com.tw/main/Main.jsp'# 搜尋頁面
    momo_mainurl ='https://www.momoshop.com.tw/'        # 拼接商品網址用 

    print('*'*15+' MOMO '+'*'*15)
    ws['a1'].value = '連線到 MOMO. . .'
    result_for_write =[]

    # - - - 啟動無頭模式 - - - 
    chrome_options = Options()
    chrome_options.add_argument('--no-sandbox') # 看這行能否解決 webdriver automation extension 打不開的問題
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    #chrome_options.add_argument('user-agent='+UserAgents[random.randint(0,len(UserAgents)-1)])
    # 加了這行的話會在  searchBox = driver.find_element_by_name('keyword') 這行報錯：   can't find this element
    # - - - 打開瀏覽器 - - -
    
    # ---- 開啟 chrome driver  -----
    if platform.system() == 'Windows': # Windows
        driver = webdriver.Chrome(os.path.split(os.path.realpath(__file__))[0]+r'\chromedriver.exe',options = chrome_options)

    elif platform.system() == 'Darwin': # Mac OS
        driver = webdriver.Chrome('./chromedriver',options = chrome_options)
    # ---- chrome driver 準備完畢 ----

    wait = WebDriverWait(driver, 5)

    try:
        # - - - 連線到 momo 前台主頁網址 - - -
        driver.get(momo_url)

        j=1
        # - - - 抓取清單內商品 - - -
        for prod in prod_list:
            ws['a1'].value = 'MOMO：{0}/{1}'.format(j,len(prod_list))
            print('搜尋商品：{} \n'.format(re.sub('___',' ',prod)))
            j+=1 
            #result[prod]['momo'] = {}
            
            # - - - 找到搜尋欄，輸入商品名，按下 Enter - - -
            searchBox = driver.find_element_by_name('keyword')
            searchBox.clear()
            searchBox.send_keys(re.sub('___',' ',prod),Keys.RETURN)

            # - - - 等搜尋結果載出來 - - -
            try:
                # pattern1: 有相符搜尋結果/ 無相符搜尋結果，給相似商品
                wait.until(EC.presence_of_element_located((By.CLASS_NAME,"totalTxt")))
            except:
                # pattern2: 連相似商品都沒有
                wait.until(EC.presence_of_element_located((By.CLASS_NAME,'newSearchFailsArea')))
                print('無搜尋結果')
                print('='*35)
                result_for_write.append(['-','-','-','-'])
                continue
                

            # - - 抓取「綜合排序」的搜尋結果 - -
            # - - - - - - 解析原始碼 - - - - - - - 
            root = bs(driver.page_source, 'html.parser')

            # - - - - 找到所有價格、品名、庫存- - - - - -
            listArea = root.find('div',class_='listArea')
            
            # print(listArea)
            prices_found  =  [x.text[1:] for x in listArea.find_all('span',class_='price')]
            names_found = [x.text for x in listArea.find_all('h3', class_='prdName')]
            urls_found = [momo_mainurl+ele['href'] for ele in listArea.select('a.goodsUrl')]
            img_urls_found = [ele['src'] for ele in listArea.select('img.prdImg')] 

            names = []
            prices = []
            urls = []
            img_urls = []

            for i in range(len(names_found)):
                if is_same_prod(prod, names_found[i], color,threeC):
                    names.append(names_found[i])
                    prices.append(prices_found[i])
                    urls.append(urls_found[i])
                    img_urls.append(img_urls_found[i])

            if names ==[]:
                print('此商品在 MOMO 無相符搜尋結果\n') 
                result_for_write.append(['-','-','-','-'])
                print('='*35)
            
            else:
                print('抓到相符商品數：',len(names),'\n')

                # 將資料存入
                # 庫存待補
                url = urls[prices.index(min(prices))]
                result_for_write.append( [min(prices), names[prices.index(min(prices))], getAmount_momo(driver, wait, url), urls[prices.index(min(prices))]] )

                #result[prod]['momo']['lowest_price'] = min(prices)
                #result[prod]['momo']['foundName'] = names[prices.index(min(prices))]
                #result[prod]['momo']['imgUrl'] = img_urls[prices.index(min(prices))]
                #result[prod]['momo']['prodUrl'] = urls[prices.index(min(prices))]
                print('='*35)

            # 等一等，當個有禮貌的爬蟲（免得被鎖IP）
            time.sleep(random.randint(1,10)*0.05)

        print('momo 抓取完畢！\n')

    except:
        logging.exception('Error occour when crawling MOMO!')
        ws['i3'].value = '未知錯誤'

    finally:
        ws['i3'].value = result_for_write
        # - - - - 結束瀏覽器，釋放記憶體空間 - - - -
        driver.quit()

def crawler_on_pchome(prod_list, ws, color, threeC):
    '''搜尋並爬取 pchome 上的商品。使用 requests。'''
    logging.info('Crawl on PChome')

    print('*'*15+' PChome '+'*'*15)
    result_for_write = []

    pc_url = 'https://ecshweb.pchome.com.tw/search/v3.3/all/results?'
    pc_mainurl='http://24h.pchome.com.tw/prod/'    # 取得商品網頁網址用
    pc_mainurl2 = 'https://d.ecimg.tw' # 取得圖片網址用

    j=0
    try:
        for prod in prod_list:
            j+=1
            ws['a1'].value = 'PChome：{0}/{1}'.format(j,len(prod_list))
            logging.info('搜尋商品：%s'%re.sub('___',' ',prod))
            print('搜尋商品：{}\n'.format(re.sub('___',' ',prod)))
            #result[prod]['PChome'] = {}
            names=[]
            prices=[]
            urls=[]
            img_urls=[]
            page=0
            
            while page<=10:
                page+=1
                try:
                    # --抓取網站回傳的 Json 格式資料 --
                    payload = {
                        'q':re.sub('___',' ',prod),
                        'page':page,
                        'sort':'rnk/dc'
                    }
                    resp = requests.get(pc_url, params=payload, headers= headers)
                    #print(resp.url)
                    resp.encoding = 'utf-8'
                    respp = resp.text
                    data = json.loads(respp)

                    if type(data)==list:
                        raise KeyError('這個頁面所抓到的 json 不是預期要抓的格式。')
                    
                    for item in data['prods']:
                        if is_same_prod(prod,item['name'], color, threeC):
                            names.append(item['name'])
                            prices.append(item['price'])
                            urls.append(pc_mainurl+item['Id'])
                            img_urls.append(pc_mainurl2+item['picB'])

                            print('查到：',item['name'])

                except KeyError:
                    print('爬到第{}頁為止'.format(page))
                    break

                except json.decoder.JSONDecodeError: # 預期：連相似結果都沒有
                    print('JSONDecodeError')
                    #print(data)
                    break

            if names ==[]:
                print('此商品在 PChome 無相符搜尋結果\n') 
                print('='*35)
                result_for_write.append(['-','-','-','-'])
                continue
                    
            else:
                print('相符商品筆數：{}\n'.format(len(names)))
                print('='*35)
                result_for_write.append([min(prices), names[prices.index(min(prices))], '稍後寫入' ,urls[prices.index(min(prices))]])
                # 將資料存入字典
                #result[prod]['PChome']['lowest_price'] = min(prices)
                #result[prod]['PChome']['foundName'] = names[prices.index(min(prices))]
                #result[prod]['PChome']['imgUrl'] = img_urls[prices.index(min(prices))]
                #result[prod]['PChome']['prodUrl'] = urls[prices.index(min(prices))]
            
            # --- 嘗試當個有禮貌的爬蟲 ---
            time.sleep(random.randint(1,10)*0.01)

    except:
        logging.exception('Error occour when crawling PChome!')

    finally:
        ws['m3'].value = result_for_write
        ws['a1'].value = ['正在查詢PChome商品可賣量']
        
        urls = [alist[-1] for alist in result_for_write]
        amounts = getAmount_PC(urls)
        print('amounts=',amounts)
        amounts_to_write = [[num] for num in amounts]
        ws['o3'].value = amounts_to_write

        print('PChome 抓取完畢！\n')
# test
#prodList = ['LAVIN 浪凡 花漾公主女性淡香水 90ml TESTER','SAMSUNG三星Galaxy A71 5G 8G/128G 6.7吋智慧手機','健司 辻利抹茶奶茶沖泡飲 22g * 30包','Jo Malone 英國梨與小蒼蘭 香水 100ml','EBI ELIE SAAB 夢幻花嫁淡香精 TESTER 90ml','MONTBLANC 萬寶龍 海洋之心女性淡香水 30ml 試用品TESTER','豐力富 紐西蘭頂級純濃奶粉 2.6 公斤']

def main():
    xw.Book.caller()
    prodList = getProdList()

    #for prod in prodList:
        #result[prod]={}

    app = xw.apps.active
    wb = xw.books.active
    ws = xw.sheets.active

    logging.info('%s Mission Start!',time.asctime())
    color = set(ws['a14'].value.split(','))
    threeC = list(ws['a18'].value.split(','))

    # test
    print(color)
    print(threeC)
    time.sleep(5)
    # test

    crawler_on_Ymall(prodList,ws,color,threeC) 
    crawler_on_Ybuy(prodList,ws,color,threeC)
    crawler_on_momo(prodList,ws,color,threeC) 
    crawler_on_pchome(prodList,ws,color,threeC) # 時有 Bug 時沒有
    crawler_on_etmall(prodList,ws,color,threeC) 
    crawler_on_FP(prodList,ws,color,threeC) # 時好時壞

    ws['a1'].value = '抓好了'
    logging.info('%s Mission Complete!',time.asctime())

# document: add the following lines at the end of your Python source file and run it.
if __name__ == '__main__':
    # log 檔案位置和 python.exe 在相同目錄下
    logging.basicConfig(filename='shopCrawlers.log', filemode='w', level=logging.DEBUG)
    xw.serve()
