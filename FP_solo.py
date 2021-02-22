# encoding = utf-8
# Author: Amy
# Last Change: 2020.11.18
# 適用系統：Windows

# version: 4.0

# Important Note:
# This ver is FOR Windows ONLY.
#  
# Dev Note:
# 1. 因為 Find Price 常常要跑第二次才會有結果，所以專門幫它做另一個按鈕
# 2. 跑程式時 .xlms 檔要開著
# 3. 此程式只能在 xlms 檔中用 RunPython 呼叫（沒有裝 mock_caller，所以不能在 python 的 console 裡呼叫這個程式）

# 待處理的問題：
# 1. 打開網址後，網頁會先經過findprice的中介頁面才進入目標網頁
# 能否用request在中介頁面抓取目標網頁，以避開用selenium解決這個問題？
# 暫用解方：selenium

# 2. 蝦皮的網站跑不出包含商品價格部分的 soup.text
# 3. momo, 樂天, PChome 24hr 這三者在無頭模式下有時會出錯，有時又不會
# 詳細測試情形見 jupyter note

from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium import webdriver

from bs4 import BeautifulSoup as bs
from difflib import SequenceMatcher
from openpyxl import load_workbook
from urllib.parse import quote
import xlwings as xw
import platform
import requests
import logging
import random
import time
import os
import re


# ===== Global Vars ======
UserAgents = [
    'Mozilla/5.0 (Macintosh; PPC Mac OS X 10_10_9 rv:2.0; he-IL) AppleWebKit/533.12.3 (KHTML, like Gecko) Version/5.0 Safari/533.12.3',
    'Mozilla/5.0 (iPad; CPU iPad OS 5_1_1 like Mac OS X) AppleWebKit/531.0 (KHTML, like Gecko) CriOS/22.0.834.0 Mobile/87E842 Safari/531.0',
    'Mozilla/5.0 (iPad; CPU iPad OS 5_1_1 like Mac OS X) AppleWebKit/531.0 (KHTML, like Gecko) CriOS/43.0.882.0 Mobile/37C008 Safari/531.0',
    'Mozilla/5.0 (iPod; U; CPU iPhone OS 3_3 like Mac OS X; uk-UA) AppleWebKit/531.31.5 (KHTML, like Gecko) Version/4.0.5 Mobile/8B117 Safari/6531.31.5',
    'Mozilla/5.0 (Macintosh; U; PPC Mac OS X 10_11_2 rv:5.0; dz-BT) AppleWebKit/532.45.2 (KHTML, like Gecko) Version/4.0 Safari/532.45.2',
    'Mozilla/5.0 (iPhone; CPU iPhone OS 6_1_6 like Mac OS X) AppleWebKit/532.2 (KHTML, like Gecko) FxiOS/10.9n4610.0 Mobile/30N900 Safari/532.2',
    'Mozilla/5.0 (iPod; U; CPU iPhone OS 3_0 like Mac OS X; gu-IN) AppleWebKit/533.20.7 (KHTML, like Gecko) Version/3.0.5 Mobile/8B115 Safari/6533.20.7']
headers = {
            'cookie': 'ECC=GoogleBot',
            # - - 使用隨機的 User-Agent - - - 
            "User-Agent":UserAgents[(random.randint(0,len(UserAgents)-1))]
            }


# ===== Sub Functions =====
def getProdList():
    '''讀取 Excel 檔取得商品名單'''
    try: # Mac
        wb = load_workbook(os.path.split(os.path.realpath(__file__))[0] +'/自動查最低價.xlsm')
    except: # windows
        wb = load_workbook('自動查最低價.xlsm')

    ws = wb.worksheets[0]

    return [cell.value for cell in list(ws.columns)[1][2:] if not cell.value==None]

def removeComma_and_toInt(str_list):
    '''移除一個由數字字串組成的 list/ str 中的每個字串中的標點符號(',','$','~','～')，並轉換字串為 int type'''
    # Last Chnaged: 2020.9.27

    # --- 處理傳入的變數並不是 list 類型的狀況 ---
    if type(str_list)==str: # str 
        astr = str_list

        while ',' in astr:
            astr = astr[:astr.find(',')]+astr[astr.find(',')+1:]

        while '$' in astr:
            astr = astr[:astr.find('$')]+astr[astr.find('$')+1:]
        
        while '~' in astr:
            astr = astr[:astr.find('~')]+astr[astr.find('~')+1:]

        while '～' in astr:
            astr = astr[:astr.find('～')] # 取左邊的

        while '元' in astr:
            astr = astr[:astr.find('元')]

        return int(astr)
        
    elif type(str_list)==int: # int
        return str_list     # 你幹嘛送不需要變換的東西進來啦
    # ---------------------------------------

    # --- 剩下的情形，預期：變數類型是 list ---
    if str_list==None:
        return None
    
    newlist = []
    if str_list==[]:
        pass

    else:
        for astr in str_list:#2020.9.1 更新
            if type(astr)==int:
                newlist.append(astr)
                continue

            while ',' in astr:
                astr = astr[:astr.find(',')]+astr[astr.find(',')+1:]

            while '$' in astr:
                astr = astr[:astr.find('$')]+astr[astr.find('$')+1:]
            
            while '~' in astr:
                astr = astr[:astr.find('~')]+astr[astr.find('~')+1:]

            while '～' in astr:
                astr = astr[:astr.find('～')] # 取左邊的

            while '元' in astr:
                astr = astr[:astr.find('元')]

            newlist.append(int(astr))    

    return newlist

def is_same_prod(prod, found, color):
    '''判斷兩個品名是否為同一商品。\n
    ＊不去特別分辨 試用品 或 非試用品'''
    logging.info('比對：\n%s\n%s',prod,found)

    # 兩品名皆轉換為小寫： 排除大小寫差異
    prod = removeComment(prod).lower()
    found = removeComment(found).lower()

    ___exist = False # default. It will be True if there is '___' in prod
    if '___' in prod:
        ___exist = True

    # 移除標點符號 
    prod = removePunc(prod)
    found = removePunc(found)

    if SequenceMatcher(None, prod, found).quick_ratio() == 1:
        logging.info('Successs: %s \n 和\n%s 吻合！',prod,found)
        return True

    elif SequenceMatcher(None, prod, found).quick_ratio() > 0.76:
        logging.info('找到相似度超過 0.76 的商品')
        logging.info('第一階段商品名處理：小寫化、移除標點符號、移除宣傳語')
        logging.info('prod=%s',prod)
        logging.info('found=%s',found)

        # 先檢查規格是否一致
        for word in color: # set color 的內容由使用者指定 (boom_data.xlsx)
            if word in prod and word not in found:
                logging.info('Failed: 指定檢查的規格不一致')
                print('指定檢查的規格不一致\n')
                return False

        # 如果商品是 3C ，用另外的特殊函式去判斷是否為相同商品 (目標是所有 3C 都在這邊處理)
        if '手機' in prod or 'apple' in prod or '平板' in prod:
            logging.info('商品和 3C 有關，套用 3C 專用規格比較法')
            return is_same_specifi(prod, found, ___exist=___exist)
        
        # 特別處理：split 後前兩個字串都不是中文 （防：英文字太多，會使中文字串的兩三字間的差異（規格）被忽略）
        if not is_chinese(prod.split()[0]) and len(prod.split())>1 and not is_chinese(prod.split()[1]):
            if SequenceMatcher(None, removeNoChinese(prod), removeNoChinese(found)).quick_ratio() > 0.75:
                s1 = ''
                s2 = ''

                # 抽掉它們不是中文的部分，重新比較
                for c in removeNoChinese(prod).split():
                    s1 +=c

                for c in removeNoChinese(found).split():
                    s2 +=c

                # 若是只拿中文的部分去互相比較相似度依然高，則再比對規格是否相符
                if SequenceMatcher(None, s1, s2).quick_ratio() > 0.75:
                    logging.info('pass: 只拿中文的部分去互相比較相似度依然高')
                    logging.info('prod_chineseOnly=%s',s1)
                    logging.info('found_chineseOnly=%s',s2)

                    pass

                else:
                    logging.info('Failed: 只拿中文的部分去互相比較，相似度不足 0.75')
                    return False

            else:
                logging.info('Failed: 相似度不足 0.75')
                return False

    # 預防：品名實際一樣，但字串相似度未超過 0.76 者
    elif SequenceMatcher(None, prod, found).quick_ratio() > 0.7 and ' ' in found:
        logging.info('第一階段商品名處理：小寫化、移除標點符號、移除宣傳語')
        logging.info('prod=%s',prod)
        logging.info('found=%s',found)
        logging.info('相似度大於 0.7 ')
        # 取走品牌名 （估計是英文），再重新比較一次
        return is_same_prod(prod, found[found.index(' ')+1:], color)
    
    # 有的時候查到的相符品名的末尾會被賣家硬塞很多關鍵字，無法通過相似度測驗
    elif len(found.split()) > 3*len(prod.split()):
        logging.info('查到品名的末尾被賣家硬塞了很多關鍵字，重新檢查')
        same_count = 0
        for word in found.split():
            if word in prod:
                same_count+=1
        
        # 如果 found 的子字串（以空格分隔）有超過 3 串都有出現在 prod 的話，接著檢查規格
        if same_count > 3:
            logging.info('pass: 子字串檢查通過')
            pass

        else:
            logging.info('Failed: same_count=%d',same_count)
            return False
        
    # 相似度太低，排除
    else:
        logging.info('Failed: 相似度太低，排除')
        return False


    # Case1: 沒有要用'___'來當作隨意數字標記
    if not ___exist:     # 未完成
        #  ---檢查品名中出現的數字和數字順序(規格)是否一致---
        # 數字：指1,2,3,...。不包含one或一二三這種。
        numIn = ''
        numFoun = ''

        for n in [num if num.isdigit() else ' ' for num in prod ]:
            numIn+=n
        
        for n in [num if num.isdigit() else ' ' for num in found ]:
            numFoun+=n

        # 如果品名中出現的 數字 與 數字順序 不一致，則判斷為不同商品
        if  numIn.split() == numFoun.split():
            # 不過濾試用品
            logging.info('Success: 規格相符\nnumIn=%s\nnumFoun=%s',numIn,numFoun)
            return True

        else:
            logging.info('numIn=%s',numIn)
            logging.info('numFoun=%s',numFoun)

            print('查到：',found)
            print('規格不符\n')

            logging.info('Failed: 品名中出現的 數字 與 數字順序 不一致，判斷為不同商品')
            return False

    # Case2: 有要用'___'來當作隨意數字標記
    else:
        #  ---檢查品名中出現的數字和數字順序(規格)是否一致: 製作正則表達式---
        # 數字：指1,2,3,...。不包含one或一二三這種。
        numIn = ''
        numFoun = ''

        for i in range(len(prod)):
            if prod[i].isdigit():
                numIn+= prod[i]

            elif i<len(prod)-2 and prod[i]=='_' and prod[i+1]=='_' and prod[i+2]:
                numIn+= '\d*'
    
        for n in found:
            if n.isdigit():
                numFoun+= n
        
        # 如果正則表達式檢查通過 (i.e., re.search() 回傳非 None 值），則回傳 True
        if re.search(numIn, numFoun):
            logging.info('Success:\n搜尋：%s\n查到：%s\n正則表達式檢查通過！',prod,found)
            return True
        
        else:
            logging.info('Failed: 正則表達式檢查未通過\nnumIn=%s\nnumFoun=%s',numIn,numFoun)
            return False
 
def removeComment(astr):
        '''用正則表達式移除對搜尋沒太大幫助的宣傳語、註記'''
        
        # unicode編碼參考： https://zh.wikipedia.org/wiki/Unicode%E5%AD%97%E7%AC%A6%E5%88%97%E8%A1%A8#%E5%9F%BA%E6%9C%AC%E6%8B%89%E4%B8%81%E5%AD%97%E6%AF%8D
        newstr_list = re.sub(r'[\[【(「（][^（「\[【(]*(新品上市|批發館|任選|效期|專案|與.+相容|免運|折後|限定|獨家\d+折|福利品|現折|限時|安裝|適用|點數[加倍]*回饋|[缺出現司櫃]貨|結帳|促銷)[^\]【）(」]*[\]】）)」]|[(]([^/ ]+/ *){1,}[^/]+[)]|效期[\W]*\d+[./]\d+[./]*\d*|\d(選|色擇)\d|.(折後.+元.|[一二兩三四五六七八九十]+色|([黑紅藍綠橙黃紫黑白金銀]/)+.|\w選\w色|只要.+起)|[^\u0020-\u0204\u4e00-\u9fa5]|[缺出現司櫃]貨[中]*|[^ ]*安裝[^ ]*|下架|[^ ]*配送',' ',astr,6).split()

        newstr = ''
        # 去除頭尾空白字元
        for word in newstr_list:
            newstr +=word + ' '
        
        return newstr[:-1]

      # 去除頭尾空白字元
        for word in newstr_list:
            newstr +=word + ' '
        
        return newstr[:-1]

def is_same_specifi(prod1, prod2, ___exist = False):
    '''判斷規格是否一致。 ＊目前只 for 手機'''
    prod1 = removePunc(removeChinese(prod1)).lower().split()
    prod2 = removePunc(removeChinese(prod2)).lower().split()

    prod1_remove_allEng = [w for w in prod1 if not w.isalpha()]
    prod2_remove_allEng = [w for w in prod2 if not w.isalpha()]

	#print(prod1_remove_allEng)
	#print(prod2_remove_allEng)

    return SequenceMatcher(None,prod1_remove_allEng,prod2_remove_allEng).quick_ratio()==1

def is_chinese(ustr):
    '''判斷一個字串是否全為中文字。 ＊只接受字串，不要亂丟int或其他什麼進來＊'''
    # 註：此處的中文字理論上包含所有繁體與簡體字
    # 只要字串包含以下任何一種，就會回傳 False ：全形標點符號、空白字元、英文、數字
    for uchar in ustr:
        if not(uchar >= u'\u4e00' and uchar <= u'\u9fa5'):
            return False

    return True

def removeNoChinese(astr):
    '''只保留原字串的中文與空白字元'''
    new_str = ''
    for c in astr:
        if is_chinese(c) or c in ' ':
            new_str+=c
        else:
            new_str+=' '

    return new_str

def removeChinese(astr):
    '''移除字串中所有中文字元'''
    new_str=''
    for c in astr:
        if not is_chinese(c):
            new_str+=c

    return new_str

def removePunc(astr):
    '''移除對商品名稱而言多餘（拿掉也不應該影響搜尋）的標點符號'''
    result = ''
    strl = list(astr)

    # 不用拿掉 + * \' .
    while set(strl).intersection({'》','《','「','」','【','】','!', '"', '#', '$', '%', '&', "'", '(', ')', ',', '-', '/', ':', ';', '<', '=', '>', '?', '@', '[', '\\', ']', '^', '`', '{', '|', '}', '~'})!=set():
        i = 0
        for c in strl:
            if c in '《》!"#$%,;<=>?@\\^`|': # 把這些符號移除
                strl.remove(c)

            elif c in '&()[]{}:-~/「」【】\'': # 把這些符號替換成空白
                strl.insert(i,' ')
                strl.remove(c)

            i+=1

    # 把新字串拼回來
    for c in strl:
        result+=c

    return result

def only_leave_alnumblank(astr):
    '''把一個字串中的數字、英文、空白符號以外的字元都移除'''
    newstr = ''

    for c in astr:
        if c.isalnum() or c.isspace():
            newstr = newstr + c
    
    return newstr

def isUrlAvailiable(driver, url, shop, price):
    '''檢查商品商店的網頁網址中的賣價和給定的賣價是否相符，抑或是該商店是否還營業著。'''
    # Last changed: 2020.10.6

    # 尚未解決的問題：
    # momo, 樂天, PChome 24hr 這三者在無頭模式下有時會出錯，有時又不會

    # 暫過 -> 一開始的 code 失敗，修改後測試皆通過
    # 尚未失敗過/ 測試全過 ->  第一個版本的 code 開始就一直表現良好（雖然不知道何時會出錯）
    # 時過時不過 -> 瓶頸。有的時候執行順利，有時卻失敗。失敗原因不明。

    driver.get(url)
    driver.implicitly_wait(5)
    soup = bs(driver.page_source,'html.parser')

    if shop == '樂天市場':# 時過時不過
        ele = soup.select_one('#auto_show_prime_price > strong > span')
        # -- 價錢位置 pattern1 --
        if ele:
            url_price = removeComma_and_toInt(ele.text)
        
        # -- 價錢位置 pattern2 --
        elif soup.select_one('#ProductMain-react-component-406930fb-6292-4b9f-b0ef-b18290b7f019 > div > div:nth-child(1) > div > section > ul > li'):
            url_price = removeComma_and_toInt(soup.select_one('#ProductMain-react-component-406930fb-6292-4b9f-b0ef-b18290b7f019 > div > div:nth-child(1) > div > section > ul > li').text)
        
        else:
            print('要查頁面',url)
            print('目前頁面：',driver.current_url)
            logging.info('目前頁面：%s',driver.current_url)
            logging.info('[樂天市場] 找不到價格欄')
            print('[樂天市場] 找不到價格欄\n')
            return False

        if url_price == price:
            return True

        else:
            #log.write('[樂天市場] 價錢和Find Price 所顯示的不符\n')
            print('Find Price 價格：',price)
            print(shop + '價格：',url_price)
            print('[樂天市場] 價錢和 Find Price 所顯示的不符\n')
            return False
                         
    elif shop =='Yahoo奇摩超級商城':
        # -- 位置 pattern1 --
        if soup.select_one('#yui_3_18_1_1_1601189656247_2212496 > div.productInfo > p > span.price'):
            url_price = removeComma_and_toInt(soup.select_one('#yui_3_18_1_1_1601189656247_2212496 > div.productInfo > p > span.price').text)

        # -- 位置 pattern2 -- (價錢會連「元」一起抓到的 pattern)
        elif soup.select_one('#ypsiif > div > div.bd.bbg.clearfix > div.right.clearfix > table > tbody > tr.webprice > td > div > span'):
            url_price = int(soup.select_one('#ypsiif > div > div.bd.bbg.clearfix > div.right.clearfix > table > tbody > tr.webprice > td > div > span').text[:-1])
        
        # -- 位置 pattern3 -- （玉山比價網頁面）
        elif soup.select_one('#yui_3_18_1_1_1601948052222_703020 > div.productInfo > p > span.price'):
            url_price = removeComma_and_toInt('#yui_3_18_1_1_1601948052222_703020 > div.productInfo > p > span.price')

        else:
            #log.write(url)
            #log.write('[Yahoo奇摩超級商城] 找不到價格欄\n')
            print('要查頁面',url)
            print('目前頁面：',driver.current_url)
            logging.info('目前頁面：%s',driver.current_url)
            logging.info('[Yahoo奇摩超級商城] 找不到價格欄')
            print('[Yahoo奇摩超級商城] 找不到價格欄\n')
            return False
        
        if url_price == price:
            return True

        else:
            #log.write('[Yahoo奇摩超級商城] 價錢和Find Price 所顯示的不符\n')
            print('Find Price 價格：',price)
            print(shop + ' 價格：',url_price)
            print('[Yahoo奇摩超級商城] 價錢和 Find Price 所顯示的不符\n')
            return  False

    elif shop =='Yahoo奇摩購物中心':
        try: 
            url_price = removeComma_and_toInt(soup.select_one('#isoredux-root > div > div.ProductItemPage__pageWrap___2CU8e > div > div:nth-child(1) > div.ProductItemPage__infoSection___3K0FH > div.ProductItemPage__rightInfoWrap___3FNQS > div > div.HeroInfo__heroInfo___1V1O8 > div > div.HeroInfo__leftWrap___3BJHV > div > div').text)
            if url_price == price:
                return True

            else:
                #log.write('[Yahoo奇摩購物中心] 價錢和Find Price 所顯示的不符\n')
                print('Find Price 價格：',price)
                print(shop + ' 價格：',url_price)
                print('[Yahoo奇摩購物中心] 價錢和 Find Price 所顯示的不符\n')
                return False

        except:
            #log.write(url)
            print('要查頁面',url)
            print('目前頁面：',driver.current_url)
            logging.info('目前頁面：%s',driver.current_url)
            if soup.select_one('#isoredux-root > div > div.ProductItemPage__pageWrap___2CU8e > div > div:nth-child(1) > div.ProductItemPage__infoSection___3K0FH > div.ProductItemPage__rightInfoWrap___3FNQS > div > div.HeroInfo__heroInfo___1V1O8 > div > div.HeroInfo__leftWrap___3BJHV > div > div')==None:
                #log.write('[Yahoo奇摩購物中心] 找不到價格欄\n')
                print('[Yahoo奇摩購物中心] 找不到價格欄\n')

            return False

    elif shop =='myfone購物':
        try: 
            url_price = removeComma_and_toInt(soup.select_one('#item-419 > div.wrapper > div.section-2 > div.prod-description > div.prod-price > span.prod-sell-price').text)

            if url_price == price:
                return True
            
            else:
                #log.write('[myfone購物]價錢和Find Price 所顯示的不符\n')
                print('Find Price 價格：',price)
                print(shop + ' 價格：',url_price)
                logging.info('Find Price 價格：%d',price)
                logging.info('%s 價格：%d',shop, url_price)
                logging.info('[myfone購物]價錢和 Find Price 所顯示的不符')
                print('[myfone購物]價錢和 Find Price 所顯示的不符\n')
                return False
            
        except:
            print('要查頁面',url)
            print('目前頁面：',driver.current_url)
            logging.info('%s',url)
            logging.info('目前頁面：%s',driver.current_url)

            if soup.select_one('#item-419 > div.wrapper > div.section-2 > div.prod-description > div.prod-price > span.prod-sell-price')==None:
                logging.info('[myfone購物] 找不到價格欄')
                print('[myfone購物] 找不到價格欄\n')
            return False
            
    elif shop == 'momo購物網':

        # -- 價格位置 pattern1 --
        if soup.select_one('#productForm > div.prdwarp > div.prdnoteArea > ul.prdPrice > li.special > span'):
            url_price = removeComma_and_toInt(soup.select_one('#productForm > div.prdwarp > div.prdnoteArea > ul.prdPrice > li.special > span').text)

        # -- 價格位置 pattern2 --
        elif soup.select_one('body > div.content > article.productPage > table > tbody > tr:nth-child(2) > td > table > tbody > tr > td.priceTxtArea > b'):
            url_price = removeComma_and_toInt(soup.select_one('body > div.content > article.productPage > table > tbody > tr:nth-child(2) > td > table > tbody > tr > td.priceTxtArea > b').text)

        # -- 銷售一空 pattern --
        elif soup.select_one('#BodyBase > div.Sold_Out > div:nth-child(1) > div > p'):
            print('[momo購物網] 售完')
            logging.info('[momo購物網] 售完')
            return False
        
        # -- 不知為何跳向了 momo主頁 pattern --
        elif driver.current_url != url:
            # 重新連線到指定網址
            driver.get(url)
            driver.implicitly_wait(2)
            soup = bs(driver.page_source,'html.parser')

            # ------ 重來 -------
            if soup.select_one('#productForm > div.prdwarp > div.prdnoteArea > ul.prdPrice > li.special > span'):
                url_price = removeComma_and_toInt(soup.select_one('#productForm > div.prdwarp > div.prdnoteArea > ul.prdPrice > li.special > span').text)

            # -- 價格位置 pattern2 --
            elif soup.select_one('body > div.content > article.productPage > table > tbody > tr:nth-child(2) > td > table > tbody > tr > td.priceTxtArea > b'):
                url_price = removeComma_and_toInt(soup.select_one('body > div.content > article.productPage > table > tbody > tr:nth-child(2) > td > table > tbody > tr > td.priceTxtArea > b').text)
            # -- 銷售一空 pattern --
            elif soup.select_one('#BodyBase > div.Sold_Out > div:nth-child(1) > div > p'):
                print('[momo購物網] 售完')
                logging.info('[momo購物網] 售完')
                return False
            
            # -- 判定為找不到 --
            else:
                print('要查頁面',url)
                print('目前頁面：',driver.current_url)
                logging.info('%s',url)
                logging.info('目前頁面：%s',driver.current_url)
                logging.info('[momo購物網] 找不到價格欄')
                print('[momo購物網] 找不到價格欄\n')
                return False

        # -- 判定為找不到 --
        else:
            print('要查頁面',url)
            print('目前頁面：',driver.current_url)
            logging.info('%s',url)
            logging.info('目前頁面：%s',driver.current_url)
            logging.info('[momo購物網] 找不到價格欄')
            print('[momo購物網] 找不到價格欄\n')
            return False
        
        if url_price == price:
            return True
        
        else:
            #log.write('[momo購物網]價錢和Find Price 所顯示的不符\n')
            logging.info('Find Price 價格：%d',price)
            logging.info('%s 價格：%d',shop, url_price)
            logging.info('[momo購物網]價錢和 Find Price 所顯示的不符')
            print('[momo購物網]價錢和 Find Price 所顯示的不符\n')
            return False  

    elif shop == 'PChome 24h購物': # 問題：抓不到網頁價錢
        #ButtonCart
        # -- 價錢位置 pattern1 --
        if soup.select_one('#PriceTotal'):
            url_price = int(soup.select_one('#PriceTotal').text)

        # -- 價錢位置 pattern2 --
        elif soup.select_one('#ProdInfo > ul.price_box > li:nth-child(1) > span > span'):
            url_price = int(soup.select_one('#ProdInfo > ul.price_box > li:nth-child(1) > span > span').text)
        # -- 偵測到完售 --
        elif soup.select_one('#ButtonCart'):
            if soup.select_one('#ButtonCart').text == '完售':
                print('[{}] 完售'.format(shop))
                logging.info('[%s] 完售',shop)
                return False

        # -- 找不到 --
        else:
            #log.write(url)
            print('要查頁面',url)
            print('目前頁面：',driver.current_url)
            logging.info('%s',url)
            logging.info('目前頁面：%s',driver.current_url)
            logging.info('[%s] 找不到價格欄',shop)
            print('[PChome 24h購物] 找不到價格欄\n')
            return False


        if url_price == price:
                return True

        else:
            #log.write('[PChome 24h購物]價錢和Find Price 所顯示的不符\n')
            print('Find Price 價格：',price)
            print(shop + ' 價格：',url_price)
            print('[PChome 24h購物]價錢和Find Price 所顯示的不符\n')
            logging.info('Find Price 價格：%d',price)
            logging.info('%s 價格：%d',shop, url_price)
            logging.info('[%s]價錢和 Find Price 所顯示的不符',shop)
            return False

    elif shop == 'ETmall東森購物網':
        # -- 抓取商品頁面價錢(用css selector定位) --
        # --- 價錢位置 pattern1 ---
        if soup.select_one('#productDetail > div:nth-child(2) > section > section > div:nth-child(3) > div.n-price__block > div.n-price__bottom > span.n-price__exlarge > span.n-price__num'):
            url_price = removeComma_and_toInt(soup.select_one('#productDetail > div:nth-child(2) > section > section > div:nth-child(3) > div.n-price__block > div.n-price__bottom > span.n-price__exlarge > span.n-price__num').text)

        # --- 價錢位置 pattern2 ---
        elif soup.select_one('#ProductDetailPrice > div > div > div > span.price-value > span'):
            url_price = removeComma_and_toInt(soup.select_one('#ProductDetailPrice > div > div > div > span.price-value > span').text)

        # --- 找不到 ---
        else:
            print(url)
            logging.info('%s',url)
            logging.info('目前頁面：%s',driver.current_url)
            logging.info('[%s] 找不到價格欄',shop)
            print('[ETmall東森購物網] 找不到價格欄\n')
            return False

        if url_price == price:
            return True

        else:
            #log.write('[ETmall東森購物網] 價錢和Find Price 所顯示的不符\n')
            print('Find Price 價格：',price)
            print(shop + ' 價格：',url_price)
            print('[ETmall東森購物網] 價錢和Find Price 所顯示的不符\n')
            logging.info('Find Price 價格：%d',price)
            logging.info('%s 價格：%d',shop, url_price)
            logging.info('[%s]價錢和 Find Price 所顯示的不符',shop)
            return False

    elif shop == '蝦皮商城':
        url_price = None
        
        # --- 按掉頁面上跳出的按鈕 ---(按「取消」)
        if soup.select_one('#modal > aside > div._2GJ_OW.undefined > div > button._3uZrWi._1UZPkA'):
            buttom = soup.select_one('#modal > aside > div._2GJ_OW.undefined > div > button._3uZrWi._1UZPkA')
            ActionChains(driver).click(buttom).perform()
            print('按按鈕')
            soup = bs(driver.page_source,'html.parser')
        
        # 猜測：會抓不到價格資訊是因為所在頁框不對
        # 因為從原始碼搜尋不到 frame 或 iframe 標籤，所以這裡採取「 loop 過所有頁框」的做法
        for handle in driver.window_handles:
            driver.switch_to_window(handle)
            soup = bs(driver.page_source,'html.parser')
            
            # -- 抓取商品頁面價錢(用css selector定位) --
            # --- 價錢位置 pattern1 ---
            if soup.select_one('#main > div > div.shopee-page-wrapper > div.page-product.page-product--mall > div.container > div.product-briefing.flex.card._2cRTS4 > div.flex.flex-auto.k-mj2F > div > div:nth-child(3) > div > div > div:nth-child(1) > div > div > div'):
                url_price = removeComma_and_toInt(soup.select_one('#main > div > div.shopee-page-wrapper > div.page-product.page-product--mall > div.container > div.product-briefing.flex.card._2cRTS4 > div.flex.flex-auto.k-mj2F > div > div:nth-child(3) > div > div > div:nth-child(1) > div > div > div').text)
                break

            # --- 價錢位置 pattern2 ---
            elif soup.select_one('#app > div > div > div.product-page.theme--ofs.nt-m > div:nth-child(4) > div.page-section.page-section--no-border.product-page__overview > div.Jvurri > div > div._1_fA5R > div.mWj_kM > div'):
                url_price = removeComma_and_toInt(soup.select_one('#app > div > div > div.product-page.theme--ofs.nt-m > div:nth-child(4) > div.page-section.page-section--no-border.product-page__overview > div.Jvurri > div > div._1_fA5R > div.mWj_kM > div').text)
                break

            # 找不到 
            else:
                print('要查頁面：',url)
                print('目前頁面：',driver.current_url)
                print('頁面text：',soup.text)
                logging.info('%s',url)
                logging.info('目前頁面：%s',driver.current_url)
                logging.info('[%s] 找不到價格欄',shop)
        
        # --- 找不到 ---
        if not url_price:
            #log.write(url)
            
            # 未預期的狀況：
            # 總是抓不到價格欄，印出 soup.text 也沒有跳出包含價格的 text

            # 猜測：frame 切換問題，抑或是從未遇到過的問題
            # 線索？
            
            logging.info('%s',url)
            logging.info('目前頁面：%s',driver.current_url)
            logging.info('[%s] 找不到價格欄',shop)
            print('[{}] 找不到價格欄\n'.format(shop))
            return False
        
        # --- 和 FP 價格相符 --> 給過
        if url_price == price:
            return True

        # --- 價格不相符 ---
        else:
            #log.write('[',shop,'] 價錢和Find Price 所顯示的不符\n')
            print('Find Price 價格：',price)
            print(shop + ' 價格：',url_price)
            print('[{}] 價錢和Find Price 所顯示的不符\n'.format(shop))
            logging.info('Find Price 價格：%d',price)
            logging.info('%s 價格：%d',shop, url_price)
            logging.info('[%s]價錢和 Find Price 所顯示的不符',shop)
            return False

    elif shop == 'Costco 好市多線上購物':
        try:
            url_price = removeComma_and_toInt(soup.select_one('#globalMessages > div.product-page-container > div.header-content-container.col-xs-12.col-sm-12.col-md-6.col-tab-6 > div:nth-child(1) > div > div.product-price > div > div > div.price-original > span > span').text)
            if url_price == price:
                return True

            else:
                #log.write('[',shop,'] 價錢和Find Price 所顯示的不符\n')
                print('Find Price 價格：',price)
                print(shop + ' 價格：',url_price)
                print('[{}] 價錢和Find Price 所顯示的不符\n'.format(shop))
                logging.info('Find Price 價格：%d',price)
                logging.info('%s 價格：%d',shop, url_price)
                logging.info('[%s]價錢和 Find Price 所顯示的不符',shop)
                return False

        except:
            #log.write(url)
            print(url)
            if soup.select_one('#globalMessages > div.product-page-container > div.header-content-container.col-xs-12.col-sm-12.col-md-6.col-tab-6 > div:nth-child(1) > div > div.product-price > div > div > div.price-original > span > span')==None:
                print('[{}] 找不到價格欄\n'.format(shop))
                logging.info('%s',url)
                logging.info('目前頁面：%s',driver.current_url)
                logging.info('[%s] 找不到價格欄',shop)

            return False
   
    elif shop == 'udn買東西':
        logging.info('尚未製作：%s',shop)
        logging.info('%s',url)
        pass

    elif shop == 'friDay購物':
        try:
            url_price = int(soup.select_one('#E3 > div > div > div.prodinfo_area > span > div.bayPricing_area > div.attract_block > span.useCash > span.price_txt').text)
            if url_price == price:
                return True
            
            else:
                #log.write('[friDay購物] 價錢和Find Price 所顯示的不符\n')
                print('Find Price 價格：',price)
                print(shop + ' 價格：',url_price)
                print('[friDay購物] 價錢和Find Price 所顯示的不符\n')
                logging.info('Find Price 價格：%d',price)
                logging.info('%s 價格：%d',shop, url_price)
                logging.info('[%s]價錢和 Find Price 所顯示的不符',shop)
                return False

        except:
            #log.write(url)
            print(url)
            if soup.select_one('#E3 > div > div > div.prodinfo_area > span > div.bayPricing_area > div.attract_block > span.useCash > span.price_txt')==None:
                logging.info('%s',url)
                logging.info('目前頁面：%s',driver.current_url)
                logging.info('[%s] 找不到價格欄',shop)
                print('[friDay購物] 找不到價格欄\n')

            return False

    elif shop == 'PChome 商店街':
        logging.info('尚未製作：%s',shop)
        logging.info('%s',url)
        pass
    
    elif shop == '家樂福線上購物網':
        logging.info('尚未製作：%s',shop)
        logging.info('%s',url)
        pass

    elif shop == 'momo摩天商城':
        try:
            url_price = int(removeComma_and_toInt(soup.select_one('#goodsForm > div.prdInnerArea > div > div.prdrightwrap > div.prdleftArea > div.prdDetailedArea > dl > dd.sellingPrice > span').text))
            if url_price == price:
                return True
            
            else:
                #log.write('[momo摩天商城] 價錢和Find Price 所顯示的不符\n')
                print('Find Price 價格：',price)
                print(shop + ' 價格：',url_price)
                print('[momo摩天商城] 價錢和Find Price 所顯示的不符\n')
                logging.info('Find Price 價格：%d',price)
                logging.info('%s 價格：%d',shop, url_price)
                logging.info('[%s]價錢和 Find Price 所顯示的不符',shop)
                return False

        except:
            #log.write(url)
            print(url)
            if soup.select_one('#goodsForm > div.prdInnerArea > div > div.prdrightwrap > div.prdleftArea > div.prdDetailedArea > dl > dd.sellingPrice > span')==None:
                logging.info('%s',url)
                logging.info('目前頁面：%s',driver.current_url)
                logging.info('[%s] 找不到價格欄',shop)
                print('[momo摩天商城] 找不到價格欄\n')
            
            return False
            
    elif shop == '森森購物網':
        logging.info('尚未製作：%s',shop)
        logging.info('%s',url)
        pass
    elif shop == '愛買線上購物':
        logging.info('尚未製作：%s',shop)
        logging.info('%s',url)
        pass
    elif shop == '大樹健康購物網':
        logging.info('尚未製作：%s',shop)
        logging.info('%s',url)
        pass
    elif shop == '小三美日平價美妝':
        logging.info('尚未製作：%s',shop)
        logging.info('%s',url)
        pass
    elif shop == '松果購物':
        logging.info('尚未製作：%s',shop)
        logging.info('%s',url)
        pass
    elif shop == '生活市集':
        logging.info('尚未製作：%s',shop)
        logging.info('%s',url)
        pass
    elif shop == 'Jollybuy 有閑':
        logging.info('尚未製作：%s',shop)
        logging.info('%s',url)
        pass
    elif shop == '熊媽媽買菜網':
        logging.info('尚未製作：%s',shop)
        logging.info('%s',url)
        pass
    elif shop == '淘寶精選':
        logging.info('尚未製作：%s',shop)
        logging.info('%s',url)
        pass

    # 如果被 pass 掉，直接回傳 True
    return True

def crawler_on_FP(prod_list, ws, color):
    '''抓取比價網站 Find Price 上的最低價'''
    fp_mainurl = 'https://www.findprice.com.tw/g/'
    fp_mainurl2 = 'https://www.findprice.com.tw'
    
    # 印分隔線
    print('*'*15+' Find Price '+'*'*15)
    result_for_write = []
    s = requests.session()

    j=0
    for prod in prod_list:
        j+=1
        ws['a1'].value = 'FP：{0}/{1}'.format(j,len(prod_list))
        print('搜尋商品：{}\n'.format(re.sub('___',' ',prod)))
        #result[prod]['FP'] = {}
        names = []
        urls = []
        prices = []
        shops = []

        url = fp_mainurl + quote(re.sub('___',' ',prod))
        req = s.get(url, headers=headers)
        count =1
        while not req.ok:
            print('連線 Find Price 失敗，請檢查網路狀態。程式接下來每隔三秒會重新嘗試連線')
            time.sleep(3)
            req = s.get(url, headers=headers)
            count+=1
            if count>5:
                raise TimeoutError('連線失敗次數過多，自動跳出 Find Price')

        req.encoding = 'utf-8'

        soup = bs(req.text ,'html.parser')

        if soup.find('div',id='GoodsGridDiv') == None or soup.find('div',id='GoodsGridDiv').a == None:
            # 有時候會出現：所有商品（包括人工搜時有搜尋結果的商品）都沒有搜尋結果的狀況
            # 原因應該是頁面抓不到這個 if 上面在判斷的 div tag
            #print(req.json) 
            #print([div.text for div in soup.find_all('div')])

            print('無任何搜尋結果')
            result_for_write.append(['-','-','-','-'])

            print('='*35)
            continue

        # 麻煩的傢伙來了
        if soup.find('div', id='HotDiv').table != None:
            link = fp_mainurl2 + soup.find('div', id='HotDiv').a['href']
            reqq = s.get(link, headers=headers)
            reqq.encoding = 'utf-8'
            soup2 = bs(reqq.text, 'html.parser')

            found_names = [tr.find_all('td')[2].a.text for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

            found_prices = [tr.find_all('td')[1].text for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

            found_urls=[tr.find_all('td')[2].a['href'] for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

            found_shops = [tr.find_all('td')[2].img['title'] for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

            found_prices = removeComma_and_toInt(found_prices)

            print('唧')

            for i in range(len(found_names)):
                if is_same_prod(prod, found_names[i], color):
                    print('查到：',found_names[i])
                    names.append(found_names[i])
                    prices.append(found_prices[i])
                    urls.append(fp_mainurl2 + found_urls[i])
                    shops.append(found_names[i])
                    
                    print('呱')


        # 一般來說會直接用這邊的程式碼
        found_names = [tr.find_all('td')[1].a.text for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')]

        found_prices = [tr.find_all('td')[1].span.text for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')]

        found_urls=[tr.find_all('td')[1].a['href'] for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')]

        found_shops = [tr.find_all('td')[1].img['title'] for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')]

        try:
            for i in range(len(found_names)):
                if is_same_prod(prod, found_names[i], color):
                    print('查到：',found_names[i])
                    names.append(found_names[i])
                    prices.append(found_prices[i])
                    urls.append(fp_mainurl2 + found_urls[i])
                    shops.append(found_shops[i])

        except AttributeError: # found_names[i] 是 list?
            print('found_names[i]=',found_names[i])
            print(type(found_names[i]))


        prices = removeComma_and_toInt(prices)

        if names == []:
            print('無相符之搜尋結果')
            print('='*35)

            result_for_write.append(['-','-','-','-'])
            continue

        print('抓到相符商品數：',len(names),'\n')
        # 存下資料以待寫入
        result_for_write.append([min(prices),names[prices.index(min(prices))],shops[prices.index(min(prices))],urls[prices.index(min(prices))]]) # 庫存待補
        print('='*35)
    
    # 將資料寫入儲存格
    ws['u3'].value = result_for_write

    print('Find Price 抓取完畢！\n')

def crawler_on_FP_new(prod_list, ws, color): # 20201006 測試：無錯誤訊息
    '''抓取比價網站 Find Price 上的最低價。新增了確認網址內容是否正確的功能。'''
    
    fp_mainurl = 'https://www.findprice.com.tw/g/'
    fp_mainurl2 = 'https://www.findprice.com.tw'

    # 印分隔線
    print('*'*15+' Find Price '+'*'*15)
    result_for_write = []
    s = requests.session()

    j=0
    for prod in prod_list:
        j+=1
        ws['a1'].value = 'FP：{0}/{1}'.format(j,len(prod_list))
        print('搜尋商品：{}\n'.format(re.sub('___',' ',prod)))
        #result[prod]['FP'] = {}
        names = []
        urls = []
        prices = []
        shops = []

        url = fp_mainurl + quote(re.sub('___',' ',prod))
        req = s.get(url, headers=headers)
        count =1
        while not req.ok:
            print('連線 Find Price 失敗，請檢查網路狀態。程式接下來每隔三秒會重新嘗試連線')
            time.sleep(3)
            req = s.get(url, headers=headers)
            count+=1
            if count>5:
                raise TimeoutError('連線失敗次數過多，自動跳出 Find Price')

        req.encoding = 'utf-8'

        soup = bs(req.text ,'html.parser')

        # ---- 檢查有無搜尋結果 ----
        if soup.find('div',id='GoodsGridDiv') == None or soup.find('div',id='GoodsGridDiv').a == None:
            # 有時候會出現：所有商品（包括人工搜時有搜尋結果的商品）都沒有搜尋結果的狀況
            # 原因應該是頁面抓不到這個 if 上面在判斷的 div tag

            for div in soup.find_all('div'):
                if '沒有符合查詢條件的商品' in div.text:
                    print('沒有符合查詢條件的商品\n')
                    break

            result_for_write.append(['-','-','-','-'])

            print('='*35)
            continue

        # 麻煩的傢伙來了
        if soup.find('div', id='HotDiv').table != None:
            link = fp_mainurl2 + soup.find('div', id='HotDiv').a['href']
            reqq = s.get(link, headers=headers)
            reqq.encoding = 'utf-8'
            soup2 = bs(reqq.text, 'html.parser')

            found_names = [tr.find_all('td')[2].a.text for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

            found_prices = [tr.find_all('td')[1].text for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

            found_urls = [tr.find_all('td')[2].a['href'] for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

            found_shops = [tr.find_all('td')[2].img['title'] for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

            found_prices = removeComma_and_toInt(found_prices)

            logging.debug('FP Crawler: the one that has to click in to crawl appear')

            for i in range(len(found_names)):
                if is_same_prod(prod, found_names[i],color):
                    logging.info('查到：%s'%found_names[i])
                    print('查到：',found_names[i])

                    while True:
                        try:
                            if isUrlAvailiable(fp_mainurl2 + found_urls[i], found_shops[i], found_prices[i]):
                                names.append(found_names[i])
                                prices.append(found_prices[i])
                                urls.append(fp_mainurl2 + found_urls[i])
                                shops.append(found_shops[i])
                                print('商品賣場檢查通過，已儲存')
                                break
                        
                        except TimeoutException:
                            print('發生錯誤，請檢查網路連線\n程式將在十秒後嘗試重新連線。')
                            time.sleep(10)
                    
                    logging.warning('FP Crawler: This log should not be print out, otherwise there is sth wrong')

        # 一般來說會直接用這邊的程式碼
        found_names = [tr.find_all('td')[1].a.text for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')]

        found_prices = removeComma_and_toInt([tr.find_all('td')[1].span.text for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')])
        found_prices = removeComma_and_toInt(found_prices)

        found_urls = [tr.find_all('td')[1].a['href'] for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')]

        found_shops = [tr.find_all('td')[1].img['title'] for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')]
        

        try:
            # ----- 準備待會要用到的 chrome driver 參數 ------
            chrome_options = Options()
            chrome_options.add_argument('user-agent='+UserAgents[random.randint(0,len(UserAgents)-1)]) # 解決抓到的原始碼是 loading 網頁載入中 的問題
            chrome_options.add_argument('--no-sandbox') # 看這行能否解決 webdriver automation extension 打不開的問題
            chrome_options.add_argument('--headless')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--enable-javascript') # 試著解決蝦皮頁面載出「please enable javascript on your browser」問題，但無明顯幫助：蝦皮頁面依然時而顯示上述訊息

            try:
                # ---- 開啟 chrome driver  -----
                if platform.system() == 'Windows': # Windows
                    driver = webdriver.Chrome(os.path.split(os.path.realpath(__file__))[0]+r'\chromedriver.exe',options = chrome_options)

                elif platform.system() == 'Darwin': # Mac OS
                    driver = webdriver.Chrome('./chromedriver',options = chrome_options)
                # ---- chrome driver 準備完畢 ----


                # --- 儲存合法商品 ---
                for i in range(len(found_names)):
                    if is_same_prod(prod, found_names[i], color):
                        logging.info('查到：%s'%found_names[i])
                        print('查到：',found_names[i])

                        if isUrlAvailiable(driver, fp_mainurl2 + found_urls[i], found_shops[i], found_prices[i]):
                            names.append(found_names[i])
                            prices.append(found_prices[i])
                            urls.append(fp_mainurl2 + found_urls[i])
                            shops.append(found_shops[i])
                            
                            print('商品賣場檢查通過，已儲存\n')

            except FileNotFoundError:
                logging.error('自動操作 chrome 失敗')
                print('自動操作 chrome 失敗')
                print('請確認 chromedriver.exe 是否已放在和此執行檔同一資料夾下')

            except TimeoutException:
                logging.error('連線逾時，請檢查網路狀態！')
                print('連線逾時，請檢查網路狀態！')
                print('請重試_')

            finally:
                logging.info('driver quit')
                driver.quit()
           

        except AttributeError: # found_names[i] 是 list?
            logging.error('AttributeError')
            logging.info('names=',names)
            logging.info('prices=',prices)
            logging.info('urls=',urls)
            logging.info('shops=',shops)

        if names == []:
            logging.info('無相符之搜尋結果')
            print('無相符之搜尋結果')
            print('='*35)

            result_for_write.append(['-','-','-','-'])
            continue
        
        logging.info('抓到相符商品數：%d'%len(names))
        print('抓到相符商品數：',len(names),'\n')

        # 存下資料以待寫入
        result_for_write.append([min(prices),names[prices.index(min(prices))],shops[prices.index(min(prices))],urls[prices.index(min(prices))]]) # 庫存待補
        print('='*35)
    
    # 將資料寫入儲存格
    ws['u3'].value = result_for_write
    logging.info('Find Price 抓取完畢！\n')
    print('Find Price 抓取完畢！\n')

def crawler_on_FP_faster(prod_list, ws, color): # 待測試
    '''抓取比價網站 Find Price 上的最低價。\n
    和 crawler_on_FP_new 的差別：會直接抓取搜尋結果中第一筆通過商店檢查的商品，而非把整頁的商品都檢查完後才輸出價格最低者。\n
    FP 的預設搜尋結果是由低價往高價排，所以這麼做應該問題不大。\n
    新增了確認網址內容是否正確的功能。'''
    
    fp_mainurl = 'https://www.findprice.com.tw/g/'
    fp_mainurl2 = 'https://www.findprice.com.tw'

    # 印分隔線
    print('*'*15+' Find Price '+'*'*15)
    result_for_write = []
    s = requests.session()

    j=0
    for prod in prod_list:
        j+=1
        ws['a1'].value = 'FP：{0}/{1}'.format(j,len(prod_list))
        print('搜尋商品：{}\n'.format(re.sub('___',' ',prod)))
        #result[prod]['FP'] = {}
        names = []
        urls = []
        prices = []
        shops = []

        url = fp_mainurl + quote(re.sub('___',' ',prod))
        req = s.get(url, headers=headers)
        count =1
        while not req.ok:
            print('連線 Find Price 失敗，請檢查網路狀態。程式接下來每隔三秒會重新嘗試連線')
            time.sleep(3)
            req = s.get(url, headers=headers)
            count+=1
            if count>5:
                raise TimeoutError('連線失敗次數過多，自動跳出 Find Price')

        req.encoding = 'utf-8'

        soup = bs(req.text ,'html.parser')

        if soup.find('div',id='GoodsGridDiv') == None or soup.find('div',id='GoodsGridDiv').a == None:
            # 有時候會出現：所有商品（包括人工搜時有搜尋結果的商品）都沒有搜尋結果的狀況
            # 原因應該是頁面抓不到這個 if 上面在判斷的 div tag

            for div in soup.find_all('div'):
                if '沒有符合查詢條件的商品' in div.text:
                    print('沒有符合查詢條件的商品\n')
                    break

            result_for_write.append(['-','-','-','-'])

            print('='*35)
            continue

        # 麻煩的傢伙來了
        if soup.find('div', id='HotDiv').table != None:
            link = fp_mainurl2 + soup.find('div', id='HotDiv').a['href']
            reqq = s.get(link, headers=headers)
            reqq.encoding = 'utf-8'
            soup2 = bs(reqq.text, 'html.parser')

            found_names = [tr.find_all('td')[2].a.text for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

            found_prices = [tr.find_all('td')[1].text for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

            found_urls = [tr.find_all('td')[2].a['href'] for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

            found_shops = [tr.find_all('td')[2].img['title'] for tr in soup2.find('div',id='GoodsGridDiv').find_all('tr')]

            found_prices = removeComma_and_toInt(found_prices)

            print('唧')

            for i in range(len(found_names)):
                if is_same_prod(prod, found_names[i],color):
                    print('查到：',found_names[i])
                    while True:
                        try:
                            if isUrlAvailiable(fp_mainurl2 + found_urls[i], found_shops[i], found_prices[i]):
                                names.append(found_names[i])
                                prices.append(found_prices[i])
                                urls.append(fp_mainurl2 + found_urls[i])
                                shops.append(found_shops[i])
                                print('商品賣場檢查通過，已儲存')
                                break
                        
                        except TimeoutException:
                            print('發生錯誤，請檢查網路連線')
                            time.sleep(10)
                    
                    print('呱')

        # 一般來說會直接用這邊的程式碼
        found_names = [tr.find_all('td')[1].a.text for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')]

        found_prices = removeComma_and_toInt([tr.find_all('td')[1].span.text for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')])
        found_prices = removeComma_and_toInt(found_prices)

        found_urls = [tr.find_all('td')[1].a['href'] for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')]

        found_shops = [tr.find_all('td')[1].img['title'] for tr in soup.find('div',id='GoodsGridDiv').find_all('tr')]

        try:
            # ----- 準備待會要用到的 chrome driver 參數 ------
            chrome_options = Options()
            chrome_options.add_argument('user-agent='+UserAgents[random.randint(0,len(UserAgents)-1)]) # 解決抓到的原始碼是 loading 網頁載入中 的問題
            chrome_options.add_argument('--no-sandbox') # 看這行能否解決 webdriver automation extension 打不開的問題
            chrome_options.add_argument('--headless')
            chrome_options.add_argument('--disable-gpu')
            chrome_options.add_argument('--enable-javascript') # 試著解決蝦皮頁面載出「please enable javascript on your browser」問題，但無明顯幫助：蝦皮頁面依然時而顯示上述訊息

            try:
                # ---- 開啟 chrome driver  -----
                if platform.system() == 'Windows': # Windows
                    driver = webdriver.Chrome(os.path.split(os.path.realpath(__file__))[0]+r'\chromedriver.exe',options = chrome_options)

                elif platform.system() == 'Darwin': # Mac OS
                    driver = webdriver.Chrome('./chromedriver',options = chrome_options)
                # ---- chrome driver 準備完畢 ----

                # --- 儲存合法商品 ---
                for i in range(len(found_names)):
                    if is_same_prod(prod, found_names[i], color):
                        print('查到：',found_names[i])
                        if isUrlAvailiable(driver, fp_mainurl2 + found_urls[i], found_shops[i], found_prices[i]):
                            names.append(found_names[i])
                            prices.append(found_prices[i])
                            urls.append(fp_mainurl2 + found_urls[i])
                            shops.append(found_shops[i])
                            
                            print('商品賣場檢查通過，已儲存\n')
                            break

            except FileNotFoundError:
                print('自動操作 chrome 失敗')
                print('請確認 chromedriver.exe 是否已放在和此執行檔同一資料夾下')

            except TimeoutException:
                print('連線逾時，請檢查網路狀態！')
                print('請重試_')

            finally:
                driver.quit()
        
        except AttributeError: # found_names[i] 是 list?
            print('AttributeError')
            print('names=',names)
            print('prices=',prices)
            print('urls=',urls)
            print('shops=',shops)

        # --- 沒有搜尋到東西 ---
        if names == []:
            print('無相符之搜尋結果')
            print('='*35)

            result_for_write.append(['-','-','-','-'])
            continue

        # 存下資料以待寫入
        result_for_write.append([min(prices),names[prices.index(min(prices))],shops[prices.index(min(prices))],urls[prices.index(min(prices))]]) # 庫存待補
        print('='*35)
    
    # 將資料寫入儲存格
    ws['u3'].value = result_for_write
    print('Find Price 抓取完畢！\n')

def main(check_shop=True):
    #log = open('#log.txt','w')
    logging.info('%s Mission Start!',time.asctime())
    
    xw.Book.caller()
    prodList = getProdList()

    app = xw.apps.active
    wb = xw.books.active
    ws = xw.sheets.active
    color = set(ws['a19'].value.split(','))

    if check_shop:
        crawler_on_FP_faster(prodList, ws, color)
    else:
        crawler_on_FP(prodList, ws, color)

    logging.info('%s Mission Complete!',time.asctime())

if __name__=='__main__':
    logging.basicConfig(filename='FP_solo.log', filemode='w', level=logging.DEBUG)
    xw.serve()